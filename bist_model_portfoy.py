#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BIST100 MODEL PORTFÖY ARACI
============================
Bilanço iyileşmesi (Katman A) + geride kalmışlık/değer (Katman B) skorlama
sistemiyle BIST100 hisselerini haftalık olarak tarar, puanlar ve Excel raporu
üretir.

VERİ KAYNAĞI
------------
Tüm veriler https://bilancoveri.com adresinin ücretsiz, anahtar gerektirmeyen
API'sinden çekilir (KAP/Borsa İstanbul kaynaklı):
  - Toplu çarpanlar: https://bilancoveri.com/api/v1/sirketler.json
  - Şirket detayı:   https://bilancoveri.com/api/v1/hisse/{ticker}.json

NASIL ÇALIŞIR
-------------
1. Toplu listeden BIST100'deki (BIST100_TICKERS) hisselerin güncel
   çarpanlarını (F/K, PD/DD, ROE, halka açıklık, fiyat) çeker.
2. Her hisse için detay uç noktasından çeyreklik bilanço/gelir tablosu
   verisini çekip:
     - Satış geliri büyümesi (YoY, aynı çeyrek karşılaştırması)
     - Net kâr büyümesi (YoY)
     - FAVÖK marjı trendi (Faaliyet Karı + Amortisman, YoY karşılaştırma)
     - Net Borç trendi (Finansal Borçlar - Nakit, YoY karşılaştırma)
   hesaplar.
3. PD/DD'yi kendi sektör ortalamasına oranlar (o an için hesaplanabilir).
4. F/K'nin kendi geçmişine göre ucuz/pahalı olduğunu VE fiyatın piyasaya
   göre göreceli performansını değerlendirebilmek için, HER ÇALIŞTIRMADA
   bir "history.csv" dosyasına o günün fiyat/F-K/PD-DD verisini ekler.
   Birkaç hafta biriktikten sonra (örn. 8+ hafta) script bu geçmişi
   kullanarak kendi F/K ortalamasına göre ucuzluk ve piyasaya göre
   göreceli getiriyi organik olarak hesaplamaya başlar. İlk haftalarda bu
   iki alt-kriter puana dahil edilmez (veri birikene kadar 0 katkı verir).
5. Risk filtresi: halka açıklık oranı ve piyasa değeri eşiklerini
   uygulayarak likidite/manipülasyon riski taşıyan hisseleri eler.
6. Sonuçları skora göre sıralayıp bir Excel dosyasına yazar
   (BIST100_Model_Portfoy_YYYYAAGG.xlsx) ve history.csv'yi günceller.

KULLANIM
--------
    python3 bist_model_portfoy.py

Her hafta aynı komutu çalıştırman yeterli. Çıktı klasörü: ./ciktilar/
Geçmiş veri dosyası: ./history.csv (bunu SİLME — zamanla skorların kalitesi
bu dosyanın büyümesiyle artar).

BIST100 LİSTESİNİ GÜNCELLEME
-----------------------------
Borsa İstanbul BIST100 endeksinin içeriğini üç ayda bir (Ocak-Nisan-
Temmuz-Ekim) gözden geçirir. Aşağıdaki BIST100_TICKERS listesi 2026
ortası itibarıyladır — üç ayda bir
https://www.borsaistanbul.com/tr/sayfa/3230/bist-pay-endeksleri
adresinden güncel listeyi kontrol edip bu listeyi güncellemen önerilir.
Alternatif olarak AUTO_TOP_N_BY_MARKET_CAP=True yaparsan script, listeyi
kullanmak yerine piyasa değerine göre ilk 100 şirketi otomatik seçer
(BIST100'ün resmi seçim kriterleriyle birebir aynı değildir ama pratik bir
yaklaşıktır).
"""

import json
import csv
import os
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime, date, timedelta
from pathlib import Path

# =============================================================================
# AYARLAR
# =============================================================================

API_BASE = "https://bilancoveri.com/api/v1"
BULK_URL = f"{API_BASE}/sirketler.json"

OUTPUT_DIR = Path("./ciktilar")
HISTORY_FILE = Path("./history.csv")
DOCS_DIR = Path(__file__).resolve().parent / "docs"   # GitHub Pages buradan yayinlanir
GIT_AUTO_PUBLISH = True    # True ise her calistirmada docs/index.html otomatik commit+push edilir

AUTO_TOP_N_BY_MARKET_CAP = False   # True yaparsan BIST100_TICKERS yok sayılır
TOP_N = 100

# Skorlama eşikleri (KAZRISK/model-portfoy sohbetinde belirlenen kurallar)
RISK_MIN_FLOAT_RATIO = 15.0        # % - bunun altı ELENİR
RISK_MIN_MARKET_CAP_MN = 5000.0    # Mn TL - cok kucuk/likit olmayan sirketleri ele (hacim verisi API'de yok, piyasa degeri proxy'si kullanildi)
RISK_MAX_NET_BORC_OZKAYNAK = 2.5   # katsayi ustu ELENIR (hesaplanabiliyorsa)

MIN_WEEKS_FOR_OWN_PE_HISTORY = 8   # bu kadar hafta biriken veriden sonra "F/K kendi ortalamasina gore" hesaplanir
MIN_WEEKS_FOR_RELATIVE_RETURN = 4  # bu kadar hafta sonra "piyasaya gore goreceli getiri" hesaplanir

REQUEST_DELAY_SEC = 0.4            # API'ye nazik davranmak icin istekler arasi bekleme
REQUEST_TIMEOUT_SEC = 20
REQUEST_RETRIES = 3

# BIST100 bilesenleri (2026 ortasi itibariyla yaklasik liste - ceyreklik guncelle!)
BIST100_TICKERS = [
    "THYAO","TUPRS","EREGL","BIMAS","SASA","KCHOL","SAHOL","GARAN","AKBNK","FROTO",
    "TOASO","PGSUS","TCELL","SISE","ASELS","ISCTR","HALKB","YKBNK","VAKBN","AKSA",
    "ARCLK","ASTOR","AEFES","AGHOL","AKSEN","ALARK","ANHYT","ANSGR","AKGRT","AKFYE",
    "AKCNS","BRSAN","BRYAT","BIOEN","BUCIM","CANTE","CCOLA","CIMSA","CWENE","DOAS",
    "DOHOL","ECILC","EKGYO","ENJSA","ENKAI","EUPWR","EUREN","FENER","GESAN","GLYHO",
    "GOLTS","GUBRF","GWIND","HEKTS","IPEKE","ISDMR","ISMEN","IZMDC","KARSN","KAYSE",
    "KLSER","KMPUR","KONTR","KONYA","KORDS","KOZAA","KOZAL","KRDMD","MAVI","MGROS",
    "MIATK","OTKAR","OYAKC","PETKM","PSGYO","QUAGR","REEDR","SAHOL","SASA","SDTTR",
    "SISE","SKBNK","SMRTG","SOKM","TAVHL","TKFEN","TSKB","TTKOM","TTRAK","TUKAS",
    "ULKER","VESBE","VESTL","YEOTK","YYLGD","ZOREN","ALTNY","BFREN","BRISA","CVKMD",
    "DSTKF","ENERY","KLYPV","KTLEV","ODAS","TABGD",
]
BIST100_TICKERS = sorted(set(BIST100_TICKERS))  # tekilleştir


# =============================================================================
# AĞ / VERİ ÇEKME
# =============================================================================

def http_get_json(url: str):
    """Basit retry'lı JSON GET. urllib kullanır (ekstra kütüphane gerekmez)."""
    last_err = None
    for attempt in range(1, REQUEST_RETRIES + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "bist-model-portfoy/1.0"})
            with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT_SEC) as resp:
                data = resp.read()
                return json.loads(data.decode("utf-8"))
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            last_err = e
            print(f"  [uyari] {url} icin deneme {attempt}/{REQUEST_RETRIES} basarisiz: {e}")
            time.sleep(1.5 * attempt)
    raise RuntimeError(f"{url} alinamadi: {last_err}")


def fetch_bulk_companies():
    print(f"Toplu carpan listesi cekiliyor: {BULK_URL}")
    payload = http_get_json(BULK_URL)
    companies = {c["ticker"]: c for c in payload.get("companies", [])}
    print(f"  -> {len(companies)} sirket alindi (generated_at={payload.get('generated_at')})")
    return companies


def fetch_company_detail(ticker: str):
    url = f"{API_BASE}/hisse/{ticker.lower()}.json"
    return http_get_json(url)


def fetch_yahoo_price_history(ticker: str, range_str: str = "2y"):
    """Yahoo Finance'in genel/ucretsiz chart API'sinden BIST hissesinin gunluk kapanis
    fiyat gecmisini ceker (gunluk/haftalik/aylik/yillik getiri hesabi icin). bilancoveri.com
    API'sinde gunluk fiyat gecmisi olmadigi icin bu ayri, anahtarsiz kaynak kullanilir.
    Basarisiz olursa bos liste doner (script'i durdurmaz)."""
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}.IS?range={range_str}&interval=1d"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT_SEC) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        result = payload["chart"]["result"][0]
        timestamps = result["timestamp"]
        closes = result["indicators"]["quote"][0]["close"]
    except Exception as e:
        print(f"  [uyari] {ticker} icin Yahoo Finance fiyat gecmisi alinamadi: {e}")
        return []

    series = [
        (datetime.utcfromtimestamp(ts).date(), c)
        for ts, c in zip(timestamps, closes) if c is not None
    ]
    series.sort(key=lambda x: x[0])
    return series


def compute_period_returns(series):
    """series: [(date, close), ...] tarihe gore artan sirali. Onceki islem gunune gore
    gunluk, ve takvim gunune gore ~1 hafta/1 ay/1 yil once ile karsilastirarak getiri (%)
    hesaplar. Yetersiz veri durumunda ilgili alan None kalir."""
    result = {"ret_1d": None, "ret_1w": None, "ret_1m": None, "ret_1y": None, "last_date": None}
    if len(series) < 2:
        return result

    latest_date, latest_close = series[-1]
    prev_date, prev_close = series[-2]
    result["last_date"] = latest_date.isoformat()
    if prev_close:
        result["ret_1d"] = (latest_close - prev_close) / prev_close * 100.0

    def find_close_on_or_before(target_date):
        best = None
        for d, c in series:
            if d <= target_date:
                best = c
            else:
                break
        return best

    for key, days_back in (("ret_1w", 7), ("ret_1m", 30), ("ret_1y", 365)):
        old_close = find_close_on_or_before(latest_date - timedelta(days=days_back))
        if old_close:
            result[key] = (latest_close - old_close) / old_close * 100.0

    return result


# =============================================================================
# FİNANSAL YARDIMCI FONKSİYONLAR
# =============================================================================

def parse_period_key(key: str):
    """'2026/6' -> (2026, 6) seklinde siralanabilir tuple doner."""
    try:
        y, q = key.split("/")
        return (int(y), int(q))
    except Exception:
        return (0, 0)


def get_field(periods: dict, period_key: str, field: str):
    p = periods.get(period_key)
    if p is None:
        return None
    v = p.get(field)
    return v


def find_latest_and_yearago(periods: dict):
    """En guncel donem anahtarini ve bir yil onceki AYNI ceyrek anahtarini bulur."""
    if not periods:
        return None, None
    keys_sorted = sorted(periods.keys(), key=parse_period_key)
    latest = keys_sorted[-1]
    y, q = parse_period_key(latest)
    yearago_key = f"{y-1}/{q}"
    if yearago_key in periods:
        return latest, yearago_key
    return latest, None


def pct_change(new, old):
    if new is None or old is None:
        return None
    try:
        if old == 0:
            return None
        return (new - old) / abs(old) * 100.0
    except Exception:
        return None


def compute_favok_proxy(periods: dict, period_key: str):
    """FAVOK yaklasik degeri = Faaliyet Kari (3DF) + Amortisman (4B).
    Ikisi de yoksa None doner."""
    op = get_field(periods, period_key, "3DF")
    dep = get_field(periods, period_key, "4B")
    if op is None:
        return None
    return op + (dep or 0.0)


def compute_net_debt(periods: dict, period_key: str):
    """Net Borc yaklasik = (Kisa+Uzun Finansal Borclar) - (Nakit ve Benzerleri).
    Bankalar/finans kuruluslari icin anlamli degildir (fin_group UFRS/XI_29K)."""
    short_debt = get_field(periods, period_key, "2AA") or 0.0
    long_debt = get_field(periods, period_key, "2BA") or 0.0
    cash = get_field(periods, period_key, "1AA") or 0.0
    return (short_debt + long_debt) - cash


def analyze_financials(detail_json: dict, fin_group: str):
    """Bir sirketin detay JSON'undan buyume/marj/borc trend metriklerini cikarir.
    Banka/finans kuruluslari (fin_group in UFRS/XI_29K) icin bircogu None doner."""
    result = {
        "revenue_growth_pct": None,
        "net_income_growth_pct": None,
        "favok_margin_trend": None,   # 'Yükseliyor' / 'Sabit' / 'Düşüyor'
        "net_debt_trend": None,       # 'Azalıyor' / 'Sabit' / 'Artıyor'
        "is_financial_sector": fin_group in ("UFRS", "XI_29K"),
    }
    periods = detail_json.get("periods", {})
    latest, yearago = find_latest_and_yearago(periods)
    if latest is None or yearago is None:
        return result

    if result["is_financial_sector"]:
        # Bankalar icin "satis" yerine faiz gelirleri (3CAC benzeri alanlar
        # bankalarda farkli kodlanir) - bu script standart XI_29 alanlarini
        # kullandigi icin bankalarda buyume hesaplamiyoruz, sadece not dusuyoruz.
        return result

    # Satis geliri buyumesi
    rev_latest = get_field(periods, latest, "3C")
    rev_yearago = get_field(periods, yearago, "3C")
    result["revenue_growth_pct"] = pct_change(rev_latest, rev_yearago)

    # Net kar buyumesi (ana ortaklik payi)
    ni_latest = get_field(periods, latest, "3Z")
    ni_yearago = get_field(periods, yearago, "3Z")
    result["net_income_growth_pct"] = pct_change(ni_latest, ni_yearago)

    # FAVOK marji trendi
    favok_latest = compute_favok_proxy(periods, latest)
    favok_yearago = compute_favok_proxy(periods, yearago)
    if favok_latest is not None and favok_yearago is not None and rev_latest and rev_yearago:
        margin_latest = favok_latest / rev_latest if rev_latest else None
        margin_yearago = favok_yearago / rev_yearago if rev_yearago else None
        if margin_latest is not None and margin_yearago is not None:
            diff = margin_latest - margin_yearago
            if diff > 0.005:      # +0.5 puan uzeri iyilesme
                result["favok_margin_trend"] = "Yükseliyor"
            elif diff < -0.005:
                result["favok_margin_trend"] = "Düşüyor"
            else:
                result["favok_margin_trend"] = "Sabit"

    # Net borc trendi (kucuk = daha iyi; azalma iyilesme demek)
    nb_latest = compute_net_debt(periods, latest)
    nb_yearago = compute_net_debt(periods, yearago)
    if nb_latest is not None and nb_yearago is not None:
        # Esik: degisimin, yil-once ozkaynagin %5'inden fazla olmasi anlamli sayilsin
        equity_yearago = get_field(periods, yearago, "2N") or 1.0
        threshold = abs(equity_yearago) * 0.05
        diff = nb_latest - nb_yearago
        if diff < -threshold:
            result["net_debt_trend"] = "Azalıyor"
        elif diff > threshold:
            result["net_debt_trend"] = "Artıyor"
        else:
            result["net_debt_trend"] = "Sabit"

    return result


# =============================================================================
# GEÇMİŞ (HISTORY) YÖNETİMİ - kendi F/K ortalaması ve göreceli getiri için
# =============================================================================

def load_history():
    """history.csv'yi {ticker: [(date, price, pe, pb), ...]} seklinde okur."""
    history = {}
    if not HISTORY_FILE.exists():
        return history
    with open(HISTORY_FILE, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            t = row["ticker"]
            history.setdefault(t, []).append({
                "date": row["date"],
                "price": float(row["price"]) if row["price"] else None,
                "pe": float(row["pe"]) if row["pe"] not in ("", "None") else None,
                "pb": float(row["pb"]) if row["pb"] not in ("", "None") else None,
            })
    return history


def append_history(companies: dict, run_date: str):
    """Bu calistirmanin fiyat/F-K/PD-DD anlik goruntusunu history.csv'ye ekler."""
    file_exists = HISTORY_FILE.exists()
    with open(HISTORY_FILE, "a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["date", "ticker", "price", "pe", "pb"])
        for ticker, c in companies.items():
            writer.writerow([run_date, ticker, c.get("price"), c.get("pe"), c.get("pb")])
    print(f"history.csv guncellendi (+{len(companies)} satir, tarih={run_date}).")


def compute_relative_return(history_for_ticker, weeks_back=4):
    """Yaklasik `weeks_back` hafta once ile karsilastirarak fiyat getirisi doner (%).
    weeks_back=None verilirse takip edilen en eski kayitla (baslangictan) karsilastirir."""
    if not history_for_ticker or len(history_for_ticker) < 2:
        return None
    sorted_hist = sorted(history_for_ticker, key=lambda r: r["date"])
    latest = sorted_hist[-1]
    if weeks_back is None:
        old = sorted_hist[0]
    else:
        target_idx = max(0, len(sorted_hist) - 1 - weeks_back)
        old = sorted_hist[target_idx]
    if latest["price"] is None or old["price"] in (None, 0):
        return None
    return (latest["price"] - old["price"]) / old["price"] * 100.0


def compute_own_pe_ratio(history_for_ticker, current_pe, min_points=8):
    """Guncel F/K'yi, gecmis history.csv'deki kendi ortalama F/K'sina oranlar."""
    if current_pe is None or current_pe <= 0:
        return None
    valid_pes = [r["pe"] for r in history_for_ticker if r.get("pe") is not None and r["pe"] > 0]
    if len(valid_pes) < min_points:
        return None
    avg_pe = sum(valid_pes) / len(valid_pes)
    if avg_pe == 0:
        return None
    return current_pe / avg_pe


# =============================================================================
# SEKTÖR ORTALAMALARI
# =============================================================================

def compute_sector_pb_averages(companies: dict):
    """Her sektor icin ortalama PD/DD hesaplar (gecerli/pozitif degerlerle)."""
    buckets = {}
    for c in companies.values():
        sec = c.get("sector_code")
        pb = c.get("pb")
        if sec and pb is not None and pb > 0:
            buckets.setdefault(sec, []).append(pb)
    return {sec: sum(vals) / len(vals) for sec, vals in buckets.items() if vals}


# =============================================================================
# SKORLAMA (KATMAN A + KATMAN B + RİSK FİLTRESİ)
# =============================================================================

def score_layer_a(fin):
    """Bilanço iyileşme skoru (max 50)."""
    if fin["is_financial_sector"]:
        return None, "Banka/finans sektörü - standart büyüme kriterleri uygulanamıyor"

    score = 0
    notes = []

    rg = fin["revenue_growth_pct"]
    if rg is not None:
        if rg >= 20:
            score += 10
        elif rg >= 10:
            score += 7
        elif rg >= 0:
            score += 4
        else:
            score -= 3
        notes.append(f"Satış büyümesi {rg:.1f}%")

    nig = fin["net_income_growth_pct"]
    if nig is not None:
        if nig >= 20:
            score += 10
        elif nig >= 10:
            score += 7
        elif nig >= 0:
            score += 4
        else:
            score -= 3
        notes.append(f"Net kâr büyümesi {nig:.1f}%")

    fm = fin["favok_margin_trend"]
    if fm is not None:
        score += {"Yükseliyor": 10, "Sabit": 4, "Düşüyor": -3}[fm]
        notes.append(f"FAVÖK marjı: {fm}")

    nd = fin["net_debt_trend"]
    if nd is not None:
        score += {"Azalıyor": 10, "Sabit": 4, "Artıyor": -3}[nd]
        notes.append(f"Net borç: {nd}")

    return score, "; ".join(notes)


def score_layer_b(company, sector_pb_avg, own_pe_ratio, relative_return):
    """Geride kalmışlık / değer skoru (max 30)."""
    score = 0
    notes = []

    if relative_return is not None:
        if relative_return <= -15:
            score += 10
        elif relative_return <= -5:
            score += 6
        elif relative_return <= 0:
            score += 3
        notes.append(f"Göreceli getiri (~history bazlı) {relative_return:.1f}%")

    if own_pe_ratio is not None:
        if own_pe_ratio < 0.7:
            score += 10
        elif own_pe_ratio < 1.0:
            score += 5
        notes.append(f"F/K ÷ kendi ortalaması = {own_pe_ratio:.2f}")

    pb = company.get("pb")
    avg_pb = sector_pb_avg
    if pb is not None and pb > 0 and avg_pb:
        ratio = pb / avg_pb
        if ratio < 0.8:
            score += 10
        elif ratio < 1.1:
            score += 5
        notes.append(f"PD/DD ÷ sektör ort. = {ratio:.2f}")

    return score, "; ".join(notes)


def apply_risk_filter(company, fin):
    """Basit risk filtresi. Hacim verisi API'de olmadigi icin piyasa degeri
    likidite proxy'si olarak kullanilir."""
    reasons = []
    float_ratio = company.get("float_ratio")
    market_cap = company.get("market_cap_mn_try")

    if float_ratio is not None and float_ratio < RISK_MIN_FLOAT_RATIO:
        reasons.append(f"Halka açıklık %{float_ratio:.1f} < %{RISK_MIN_FLOAT_RATIO}")
    if market_cap is not None and market_cap < RISK_MIN_MARKET_CAP_MN:
        reasons.append(f"Piyasa değeri {market_cap:.0f} Mn TL < {RISK_MIN_MARKET_CAP_MN:.0f} Mn TL")

    passed = len(reasons) == 0
    return passed, "; ".join(reasons) if reasons else "GEÇTİ"


# =============================================================================
# EXCEL RAPORU
# =============================================================================

def write_excel_report(rows, run_date_str, output_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Portföy"
    ws.sheet_view.showGridLines = False

    title_font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    header_font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    normal_font = Font(name=FONT, size=10)
    warn_font = Font(name=FONT, size=9, bold=True, color="FFFFFF")

    fill_navy = PatternFill("solid", fgColor="1F3864")
    fill_red = PatternFill("solid", fgColor="C00000")
    fill_header = PatternFill("solid", fgColor="2E75B6")
    fill_calc = PatternFill("solid", fgColor="E2EFDA")

    thin = Side(style="thin", color="BFBFBF")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:Q1")
    ws["A1"] = f"BIST100 MODEL PORTFÖY — Haftalık Tarama ({run_date_str})"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:Q2")
    ws["A2"] = ("UYARI: Yatırım tavsiyesi değildir, sistematik bir tarama aracıdır. "
                "Veri kaynağı: bilancoveri.com (KAP/BIST) ve Yahoo Finance (fiyat getirisi). Karar sizindir.")
    ws["A2"].font = warn_font
    ws["A2"].fill = fill_red
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 20

    headers = ["Sıra", "Kod", "Şirket", "Sektör", "Toplam Skor", "Katman A",
               "Katman B", "Risk Filtresi", "F/K", "PD/DD", "ROE %",
               "Günlük %", "Haftalık %", "Aylık %", "Yıllık %",
               "Katman A Notları", "Katman B Notları"]
    widths = [6, 8, 26, 16, 12, 10, 10, 22, 8, 8, 8, 10, 10, 10, 10, 40, 36]
    for i, (h, w) in enumerate(zip(headers, widths)):
        col = get_column_letter(i + 1)
        ws.column_dimensions[col].width = w
        c = ws[f"{col}4"]
        c.value = h
        c.font = header_font
        c.fill = fill_header
        c.alignment = center
        c.border = border_all
    ws.row_dimensions[4].height = 30

    # Sadece skorlanabilenleri (Katman A hesaplanabilen ve elenmeyenleri) skora göre sırala,
    # skorlanamayanlar ve elenenler altta listelensin.
    def sort_key(r):
        if not r["risk_passed"]:
            return (2, 0)
        if r["total_score"] is None:
            return (1, 0)
        return (0, -r["total_score"])

    rows_sorted = sorted(rows, key=sort_key)

    return_cols = (12, 13, 14, 15)
    r_idx = 5
    for i, r in enumerate(rows_sorted):
        ws.cell(row=r_idx, column=1, value=i + 1)
        ws.cell(row=r_idx, column=2, value=r["ticker"])
        ws.cell(row=r_idx, column=3, value=r["name"])
        ws.cell(row=r_idx, column=4, value=r["sector"])
        ws.cell(row=r_idx, column=5, value=r["total_score"] if r["risk_passed"] else 0)
        ws.cell(row=r_idx, column=6, value=r["layer_a_score"])
        ws.cell(row=r_idx, column=7, value=r["layer_b_score"])
        ws.cell(row=r_idx, column=8, value=r["risk_status"])
        ws.cell(row=r_idx, column=9, value=r["pe"])
        ws.cell(row=r_idx, column=10, value=r["pb"])
        ws.cell(row=r_idx, column=11, value=r["roe"])
        ws.cell(row=r_idx, column=12, value=r["ret_1d"])
        ws.cell(row=r_idx, column=13, value=r["ret_1w"])
        ws.cell(row=r_idx, column=14, value=r["ret_1m"])
        ws.cell(row=r_idx, column=15, value=r["ret_1y"])
        ws.cell(row=r_idx, column=16, value=r["layer_a_notes"])
        ws.cell(row=r_idx, column=17, value=r["layer_b_notes"])
        for col in range(1, 18):
            cell = ws.cell(row=r_idx, column=col)
            cell.font = normal_font
            cell.border = border_all
            cell.alignment = left if col in (3, 16, 17) else center
            if col in return_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"
        r_idx += 1

    last_row = r_idx - 1
    if last_row >= 5:
        ws.conditional_formatting.add(
            f"E5:E{last_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["55"],
                       fill=PatternFill("solid", fgColor="8EA9DB")))
        ws.conditional_formatting.add(
            f"E5:E{last_row}",
            CellIsRule(operator="between", formula=["30", "54"],
                       fill=PatternFill("solid", fgColor="A9D18E")))
        ws.conditional_formatting.add(
            f"H5:H{last_row}",
            CellIsRule(operator="notEqual", formula=['"GEÇTİ"'],
                       fill=PatternFill("solid", fgColor="FF7C80")))
        for col_letter in ("L", "M", "N", "O"):
            ws.conditional_formatting.add(
                f"{col_letter}5:{col_letter}{last_row}",
                CellIsRule(operator="greaterThan", formula=["0"],
                           fill=PatternFill("solid", fgColor="A9D18E")))
            ws.conditional_formatting.add(
                f"{col_letter}5:{col_letter}{last_row}",
                CellIsRule(operator="lessThan", formula=["0"],
                           fill=PatternFill("solid", fgColor="FF7C80")))

    ws.freeze_panes = "A5"

    note_row = last_row + 3
    ws.merge_cells(f"A{note_row}:Q{note_row}")
    ws[f"A{note_row}"] = ("Not: Katman B'deki 'göreceli getiri' ve 'kendi F/K ortalaması' alt-kriterleri "
                           f"history.csv dosyasındaki birikmiş veriye bağlıdır — ilk {MIN_WEEKS_FOR_RELATIVE_RETURN}-"
                           f"{MIN_WEEKS_FOR_OWN_PE_HISTORY} hafta boyunca eksik/0 gelebilir, bu normaldir. "
                           "Günlük/Haftalık/Aylık/Yıllık % sütunları Yahoo Finance'ten çekilen gerçek fiyat "
                           "geçmişine dayanır; bir hissenin verisi çekilemezse boş kalır.")
    ws[f"A{note_row}"].font = Font(name=FONT, size=9, italic=True, color="7F7F7F")
    ws[f"A{note_row}"].alignment = left

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel raporu yazildi: {output_path}")


# =============================================================================
# HTML PANO (bagimsiz, tek dosyalik interaktif rapor)
# =============================================================================

DASHBOARD_TEMPLATE = """<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>BIST100 Model Portföy</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500;600&display=swap" rel="stylesheet">

<style>
  :root {
    --bg: #EEF2F3;
    --bg-elevated: #FFFFFF;
    --bg-sunken: #E4EAEC;
    --ink: #101B23;
    --ink-muted: #55666F;
    --ink-faint: #8598A0;
    --border: #D5DEE1;
    --border-strong: #BCC9CD;
    --accent: #A06B1F;
    --accent-ink: #FFFFFF;
    --accent-soft: rgba(160, 107, 31, 0.12);
    --positive: #1D7A54;
    --positive-bg: rgba(29, 122, 84, 0.11);
    --negative: #B5402C;
    --negative-bg: rgba(181, 64, 44, 0.11);
    --tier-strong-bg: rgba(160, 107, 31, 0.14);
    --tier-strong-ink: #8A5A18;
    --tier-watch-bg: rgba(29, 122, 84, 0.10);
    --tier-watch-ink: #1D7A54;
    --tier-weak-bg: rgba(85, 102, 111, 0.08);
    --tier-weak-ink: #55666F;
    --tier-out-bg: rgba(181, 64, 44, 0.09);
    --tier-out-ink: #B5402C;
    --shadow: 0 1px 2px rgba(16, 27, 35, 0.06), 0 8px 24px -12px rgba(16, 27, 35, 0.18);
    --font-display: 'Fraunces', Georgia, serif;
    --font-body: 'IBM Plex Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    --font-mono: 'IBM Plex Mono', 'SFMono-Regular', Consolas, monospace;
  }

  @media (prefers-color-scheme: dark) {
    :root:not([data-theme="light"]) {
      --bg: #0C131A;
      --bg-elevated: #131C24;
      --bg-sunken: #0A1017;
      --ink: #E8EEF0;
      --ink-muted: #93A5AD;
      --ink-faint: #607079;
      --border: #253440;
      --border-strong: #33454F;
      --accent: #D9A24E;
      --accent-ink: #1A1206;
      --accent-soft: rgba(217, 162, 78, 0.14);
      --positive: #4FC489;
      --positive-bg: rgba(79, 196, 137, 0.13);
      --negative: #E37363;
      --negative-bg: rgba(227, 115, 99, 0.13);
      --tier-strong-bg: rgba(217, 162, 78, 0.16);
      --tier-strong-ink: #E8B968;
      --tier-watch-bg: rgba(79, 196, 137, 0.13);
      --tier-watch-ink: #4FC489;
      --tier-weak-bg: rgba(147, 165, 173, 0.10);
      --tier-weak-ink: #93A5AD;
      --tier-out-bg: rgba(227, 115, 99, 0.12);
      --tier-out-ink: #E37363;
      --shadow: 0 1px 2px rgba(0, 0, 0, 0.3), 0 8px 24px -12px rgba(0, 0, 0, 0.5);
    }
  }

  :root[data-theme="dark"] {
    --bg: #0C131A;
    --bg-elevated: #131C24;
    --bg-sunken: #0A1017;
    --ink: #E8EEF0;
    --ink-muted: #93A5AD;
    --ink-faint: #607079;
    --border: #253440;
    --border-strong: #33454F;
    --accent: #D9A24E;
    --accent-ink: #1A1206;
    --accent-soft: rgba(217, 162, 78, 0.14);
    --positive: #4FC489;
    --positive-bg: rgba(79, 196, 137, 0.13);
    --negative: #E37363;
    --negative-bg: rgba(227, 115, 99, 0.13);
    --tier-strong-bg: rgba(217, 162, 78, 0.16);
    --tier-strong-ink: #E8B968;
    --tier-watch-bg: rgba(79, 196, 137, 0.13);
    --tier-watch-ink: #4FC489;
    --tier-weak-bg: rgba(147, 165, 173, 0.10);
    --tier-weak-ink: #93A5AD;
    --tier-out-bg: rgba(227, 115, 99, 0.12);
    --tier-out-ink: #E37363;
    --shadow: 0 1px 2px rgba(0, 0, 0, 0.3), 0 8px 24px -12px rgba(0, 0, 0, 0.5);
  }

  * { box-sizing: border-box; }
  html, body { height: 100%; }
  body {
    margin: 0;
    background: var(--bg);
    color: var(--ink);
    font-family: var(--font-body);
    font-size: 15px;
    line-height: 1.45;
    -webkit-font-smoothing: antialiased;
  }
  @media (prefers-reduced-motion: reduce) {
    * { animation-duration: 0.001ms !important; transition-duration: 0.001ms !important; }
  }

  .app {
    display: flex;
    flex-direction: column;
    height: 100vh;
    max-width: 1400px;
    margin: 0 auto;
    padding: 0 clamp(14px, 3vw, 32px);
  }

  /* ---- Masthead ---- */
  .masthead {
    display: flex;
    flex-wrap: wrap;
    align-items: baseline;
    justify-content: space-between;
    gap: 8px 24px;
    padding: 22px 2px 16px;
    border-bottom: 1px solid var(--border);
  }
  .masthead-id { display: flex; flex-direction: column; gap: 2px; }
  .eyebrow {
    font-family: var(--font-mono);
    font-size: 11px;
    font-weight: 500;
    letter-spacing: 0.11em;
    text-transform: uppercase;
    color: var(--accent);
  }
  h1 {
    margin: 0;
    font-family: var(--font-display);
    font-weight: 600;
    font-size: clamp(22px, 3vw, 30px);
    letter-spacing: -0.01em;
    text-wrap: balance;
  }
  .masthead-meta {
    text-align: right;
    font-size: 12.5px;
    color: var(--ink-muted);
    line-height: 1.5;
  }
  .masthead-meta strong { color: var(--ink); font-weight: 600; }

  /* ---- Stat tiles ---- */
  .stats {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(128px, 1fr));
    gap: 1px;
    background: var(--border);
    border: 1px solid var(--border);
    border-radius: 10px;
    overflow: hidden;
    margin: 14px 0;
    box-shadow: var(--shadow);
  }
  .stat {
    background: var(--bg-elevated);
    padding: 12px 16px;
    display: flex;
    flex-direction: column;
    gap: 3px;
  }
  .stat-label {
    font-size: 10.5px;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: var(--ink-faint);
    font-weight: 600;
  }
  .stat-value {
    font-family: var(--font-mono);
    font-size: 21px;
    font-weight: 600;
    font-variant-numeric: tabular-nums;
    color: var(--ink);
  }
  .stat-value.pos { color: var(--positive); }
  .stat-value.neg { color: var(--negative); }
  .stat-sub { font-size: 11px; color: var(--ink-muted); }

  /* ---- Toolbar ---- */
  .toolbar {
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    gap: 10px;
    padding-bottom: 12px;
  }
  .search-field {
    position: relative;
    flex: 1 1 220px;
    min-width: 160px;
  }
  .search-field input {
    width: 100%;
    padding: 8px 12px 8px 32px;
    border-radius: 8px;
    border: 1px solid var(--border-strong);
    background: var(--bg-elevated);
    color: var(--ink);
    font-family: var(--font-body);
    font-size: 13.5px;
  }
  .search-field input:focus-visible { outline: 2px solid var(--accent); outline-offset: 1px; }
  .search-field::before {
    content: "";
    position: absolute;
    left: 11px; top: 50%;
    width: 12px; height: 12px;
    margin-top: -6px;
    border: 1.5px solid var(--ink-faint);
    border-radius: 50%;
    box-shadow: 5px 5px 0 -3px var(--ink-faint);
  }
  select, .toggle-chip {
    padding: 8px 12px;
    border-radius: 8px;
    border: 1px solid var(--border-strong);
    background: var(--bg-elevated);
    color: var(--ink);
    font-family: var(--font-body);
    font-size: 13px;
    cursor: pointer;
  }
  select:focus-visible, .toggle-chip:focus-visible { outline: 2px solid var(--accent); outline-offset: 1px; }
  .toggle-chip {
    display: inline-flex;
    align-items: center;
    gap: 7px;
    user-select: none;
    white-space: nowrap;
  }
  .toggle-chip input { accent-color: var(--accent); width: 14px; height: 14px; }
  .result-count {
    margin-left: auto;
    font-size: 12px;
    color: var(--ink-muted);
    font-variant-numeric: tabular-nums;
  }

  /* ---- Table ---- */
  .table-wrap {
    flex: 1;
    min-height: 0;
    overflow: auto;
    border: 1px solid var(--border);
    border-radius: 10px;
    background: var(--bg-elevated);
    box-shadow: var(--shadow);
  }
  table { border-collapse: collapse; width: 100%; min-width: 980px; }
  thead th {
    position: sticky; top: 0; z-index: 2;
    background: var(--bg-sunken);
    border-bottom: 1px solid var(--border-strong);
    text-align: left;
    padding: 9px 12px;
    font-size: 10.5px;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--ink-muted);
    font-weight: 600;
    white-space: nowrap;
    cursor: pointer;
  }
  thead th:hover { color: var(--ink); }
  thead th.active { color: var(--accent); }
  thead th .arrow { font-size: 9px; margin-left: 3px; opacity: 0.8; }
  thead th.num, td.num { text-align: right; }
  tbody tr.row { cursor: pointer; border-bottom: 1px solid var(--border); }
  tbody tr.row:hover { background: var(--accent-soft); }
  tbody tr.row.out { opacity: 0.55; }
  tbody tr.row.expanded { background: var(--accent-soft); }
  td { padding: 8px 12px; font-size: 13px; white-space: nowrap; }
  td.ticker { font-family: var(--font-mono); font-weight: 600; letter-spacing: 0.01em; }
  td.name { white-space: normal; min-width: 180px; }
  td.name .sector { display: block; font-size: 11px; color: var(--ink-faint); font-weight: 400; margin-top: 1px; }
  td.num { font-family: var(--font-mono); font-variant-numeric: tabular-nums; color: var(--ink-muted); }
  .ret.pos { color: var(--positive); }
  .ret.neg { color: var(--negative); }
  .ret.zero { color: var(--ink-faint); }

  .score-cell { display: flex; align-items: center; gap: 8px; }
  .score-num { font-family: var(--font-mono); font-weight: 600; font-variant-numeric: tabular-nums; width: 26px; text-align: right; }
  .score-track { width: 46px; height: 6px; border-radius: 4px; background: var(--bg-sunken); overflow: hidden; flex-shrink: 0; }
  .score-fill { height: 100%; border-radius: 4px; background: var(--ink-faint); }
  .row.strong .score-fill, .row.strong .score-num { background: var(--tier-strong-ink); color: var(--tier-strong-ink); }
  .row.strong .score-fill { background: var(--tier-strong-ink); }
  .row.watch .score-fill { background: var(--tier-watch-ink); }
  .row.watch .score-num { color: var(--tier-watch-ink); }
  .row.strong .score-num { color: var(--tier-strong-ink); }
  .row.weak .score-fill { background: var(--ink-faint); }
  .row.out .score-fill { background: var(--tier-out-ink); }

  .chip {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 20px;
    font-size: 10.5px;
    font-weight: 600;
    letter-spacing: 0.02em;
    white-space: nowrap;
  }
  .chip.strong { background: var(--tier-strong-bg); color: var(--tier-strong-ink); }
  .chip.watch { background: var(--tier-watch-bg); color: var(--tier-watch-ink); }
  .chip.weak { background: var(--tier-weak-bg); color: var(--tier-weak-ink); }
  .chip.out { background: var(--tier-out-bg); color: var(--tier-out-ink); }

  td.chevron { width: 18px; color: var(--ink-faint); transition: transform 0.15s ease; }
  tr.row.expanded td.chevron { transform: rotate(90deg); color: var(--accent); }

  tr.detail td {
    white-space: normal;
    background: var(--bg-sunken);
    padding: 0;
    border-bottom: 1px solid var(--border);
  }
  .detail-inner {
    display: none;
    grid-template-columns: 1fr 1fr auto;
    gap: 18px;
    padding: 14px 18px;
  }
  tr.detail.open .detail-inner { display: grid; }
  @media (max-width: 760px) {
    .detail-inner { grid-template-columns: 1fr; }
  }
  .detail-block h4 {
    margin: 0 0 4px;
    font-size: 10.5px;
    text-transform: uppercase;
    letter-spacing: 0.07em;
    color: var(--ink-faint);
    font-weight: 600;
  }
  .detail-block p { margin: 0; font-size: 13px; color: var(--ink-muted); line-height: 1.5; }
  .detail-risk { min-width: 160px; }
  .detail-risk .risk-line { font-size: 12.5px; color: var(--ink-muted); margin-top: 4px; }

  @media (max-width: 760px) {
    .app { height: auto; min-height: 100vh; }
    .table-wrap { flex: none; }
  }

  .empty-state {
    padding: 48px 20px;
    text-align: center;
    color: var(--ink-muted);
    font-size: 13.5px;
  }

  /* ---- Footer ---- */
  footer {
    padding: 10px 2px 16px;
    font-size: 11.5px;
    color: var(--ink-faint);
    border-top: 1px solid var(--border);
    display: flex;
    flex-wrap: wrap;
    justify-content: space-between;
    gap: 6px 20px;
  }
  footer a { color: inherit; }

  ::-webkit-scrollbar { width: 10px; height: 10px; }
  ::-webkit-scrollbar-track { background: transparent; }
  ::-webkit-scrollbar-thumb { background: var(--border-strong); border-radius: 6px; }
</style>
</head>
<body>
<div class="app">
  <header class="masthead">
    <div class="masthead-id">
      <span class="eyebrow">BIST100 · Haftalık Tarama</span>
      <h1>Model Portföy</h1>
    </div>
    <div class="masthead-meta">
      Tarama tarihi: <strong id="run-date">—</strong><br>
      Kaynak: bilancoveri.com (bilanço) · Yahoo Finance (getiri)
    </div>
  </header>

  <section class="stats" id="stats"></section>

  <section class="toolbar">
    <div class="search-field">
      <input id="search" type="text" placeholder="Kod veya şirket ara…" autocomplete="off">
    </div>
    <select id="sector-filter"></select>
    <label class="toggle-chip">
      <input type="checkbox" id="risk-toggle" checked>
      Sadece riski geçenler
    </label>
    <span class="result-count" id="result-count"></span>
  </section>

  <div class="table-wrap">
    <table>
      <thead>
        <tr id="head-row"></tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    <div class="empty-state" id="empty-state" style="display:none;">Aramanla eşleşen hisse yok.</div>
  </div>

  <footer>
    <span>Yatırım tavsiyesi değildir — sistematik bir tarama aracıdır. Karar sizindir.</span>
    <span id="stock-count-footer"></span>
  </footer>
</div>

<script>
  const DATA = __BIST_DATA__;
  const RUN_DATE = "__RUN_DATE__";

  document.getElementById('run-date').textContent = RUN_DATE;
  document.getElementById('stock-count-footer').textContent = DATA.length + ' hisse tarandı';

  DATA.forEach(d => { d.risk_passed = d.risk_status === 'GEÇTİ'; });

  function tierOf(d) {
    if (!d.risk_passed) return 'out';
    if (d.total_score >= 55) return 'strong';
    if (d.total_score >= 30) return 'watch';
    return 'weak';
  }
  const TIER_LABEL = { strong: 'Öne çıkan', watch: 'İzlemede', weak: 'Zayıf', out: 'Elendi' };

  const scores = DATA.map(d => d.total_score).filter(v => v != null);
  const scoreMin = Math.min(0, ...scores);
  const scoreMax = Math.max(10, ...scores);

  function fmtPct(v) {
    if (v == null) return '<span class="ret zero">—</span>';
    const cls = v > 0.05 ? 'pos' : v < -0.05 ? 'neg' : 'zero';
    const sign = v > 0 ? '+' : '';
    return `<span class="ret ${cls}">${sign}${v.toFixed(1)}%</span>`;
  }
  function fmtNum(v, digits) {
    if (v == null) return '—';
    return v.toFixed(digits == null ? 2 : digits);
  }

  const COLUMNS = [
    { key: 'ticker', label: 'Kod', sort: (d) => d.ticker },
    { key: 'name', label: 'Şirket', sort: (d) => d.name },
    { key: 'total_score', label: 'Skor', num: true, sort: (d) => d.total_score ?? -999 },
    { key: 'pe', label: 'F/K', num: true, sort: (d) => d.pe ?? -Infinity },
    { key: 'pb', label: 'PD/DD', num: true, sort: (d) => d.pb ?? -Infinity },
    { key: 'roe', label: 'ROE %', num: true, sort: (d) => d.roe ?? -Infinity },
    { key: 'ret_1d', label: 'Günlük', num: true, sort: (d) => d.ret_1d ?? -Infinity },
    { key: 'ret_1w', label: 'Haftalık', num: true, sort: (d) => d.ret_1w ?? -Infinity },
    { key: 'ret_1m', label: 'Aylık', num: true, sort: (d) => d.ret_1m ?? -Infinity },
    { key: 'ret_1y', label: 'Yıllık', num: true, sort: (d) => d.ret_1y ?? -Infinity },
  ];

  let sortKey = 'total_score';
  let sortDir = -1;
  let expandedTicker = null;

  const headRow = document.getElementById('head-row');
  headRow.innerHTML = '<th style="width:18px;"></th>' + COLUMNS.map(c =>
    `<th data-key="${c.key}" class="${c.num ? 'num' : ''}">${c.label}<span class="arrow"></span></th>`
  ).join('');

  headRow.querySelectorAll('th[data-key]').forEach(th => {
    th.addEventListener('click', () => {
      const key = th.dataset.key;
      if (sortKey === key) { sortDir *= -1; } else { sortKey = key; sortDir = key === 'ticker' || key === 'name' ? 1 : -1; }
      render();
    });
  });

  const sectorSel = document.getElementById('sector-filter');
  const sectors = [...new Set(DATA.map(d => d.sector).filter(Boolean))].sort((a, b) => a.localeCompare(b, 'tr'));
  sectorSel.innerHTML = '<option value="">Tüm sektörler</option>' + sectors.map(s => `<option value="${s}">${s}</option>`).join('');

  const searchInput = document.getElementById('search');
  const riskToggle = document.getElementById('risk-toggle');
  const tbody = document.getElementById('tbody');
  const emptyState = document.getElementById('empty-state');
  const resultCount = document.getElementById('result-count');

  function renderStats() {
    const passed = DATA.filter(d => d.risk_passed);
    const strong = passed.filter(d => d.total_score >= 55);
    const watch = passed.filter(d => d.total_score >= 30 && d.total_score < 55);
    const withYear = DATA.filter(d => d.ret_1y != null);
    const avgYear = withYear.length ? withYear.reduce((a, d) => a + d.ret_1y, 0) / withYear.length : null;
    const top = [...passed].sort((a, b) => b.total_score - a.total_score)[0];

    const tiles = [
      { label: 'Taranan Hisse', value: DATA.length, sub: 'BIST100' },
      { label: 'Riski Geçen', value: passed.length, sub: (DATA.length - passed.length) + ' elendi' },
      { label: 'Öne Çıkan', value: strong.length, sub: 'skor ≥ 55' },
      { label: 'İzlemede', value: watch.length, sub: 'skor 30–54' },
      { label: 'Ort. Yıllık Getiri', value: avgYear == null ? '—' : (avgYear > 0 ? '+' : '') + avgYear.toFixed(1) + '%', sub: 'Yahoo Finance', cls: avgYear > 0 ? 'pos' : avgYear < 0 ? 'neg' : '' },
      { label: 'En Yüksek Skor', value: top ? top.ticker : '—', sub: top ? top.total_score + ' puan' : '' },
    ];
    document.getElementById('stats').innerHTML = tiles.map(t =>
      `<div class="stat"><span class="stat-label">${t.label}</span><span class="stat-value ${t.cls || ''}">${t.value}</span><span class="stat-sub">${t.sub}</span></div>`
    ).join('');
  }

  function currentFilter() {
    const q = searchInput.value.trim().toLocaleLowerCase('tr');
    const sector = sectorSel.value;
    const onlyPassed = riskToggle.checked;
    return DATA.filter(d => {
      if (onlyPassed && !d.risk_passed) return false;
      if (sector && d.sector !== sector) return false;
      if (q && !(d.ticker.toLocaleLowerCase('tr').includes(q) || (d.name || '').toLocaleLowerCase('tr').includes(q))) return false;
      return true;
    });
  }

  function render() {
    headRow.querySelectorAll('th[data-key]').forEach(th => {
      th.classList.toggle('active', th.dataset.key === sortKey);
      th.querySelector('.arrow').textContent = th.dataset.key === sortKey ? (sortDir === 1 ? '▲' : '▼') : '';
    });

    const col = COLUMNS.find(c => c.key === sortKey);
    const rows = currentFilter().sort((a, b) => {
      const av = col.sort(a), bv = col.sort(b);
      if (typeof av === 'string') return av.localeCompare(bv, 'tr') * sortDir;
      return (av - bv) * sortDir;
    });

    resultCount.textContent = rows.length + ' / ' + DATA.length + ' hisse';
    emptyState.style.display = rows.length ? 'none' : 'block';

    tbody.innerHTML = rows.map(d => {
      const tier = tierOf(d);
      const barPct = Math.max(0, Math.min(100, ((d.total_score - scoreMin) / (scoreMax - scoreMin)) * 100));
      const isOpen = d.ticker === expandedTicker;
      return `
      <tr class="row ${tier} ${isOpen ? 'expanded' : ''}" data-ticker="${d.ticker}">
        <td class="chevron">›</td>
        <td class="ticker">${d.ticker}</td>
        <td class="name">${d.name || ''}<span class="sector">${d.sector || ''}</span></td>
        <td class="num"><div class="score-cell"><span class="score-num">${d.total_score ?? '—'}</span><span class="score-track"><span class="score-fill" style="width:${barPct}%"></span></span></div></td>
        <td class="num">${fmtNum(d.pe, 1)}</td>
        <td class="num">${fmtNum(d.pb, 2)}</td>
        <td class="num">${fmtNum(d.roe, 1)}</td>
        <td class="num">${fmtPct(d.ret_1d)}</td>
        <td class="num">${fmtPct(d.ret_1w)}</td>
        <td class="num">${fmtPct(d.ret_1m)}</td>
        <td class="num">${fmtPct(d.ret_1y)}</td>
      </tr>
      <tr class="detail ${isOpen ? 'open' : ''}" data-ticker-detail="${d.ticker}">
        <td colspan="${COLUMNS.length + 1}">
          <div class="detail-inner">
            <div class="detail-block">
              <h4>Katman A · Bilanço İyileşmesi (${d.layer_a_score ?? '—'} puan)</h4>
              <p>${d.layer_a_notes || 'Veri yetersiz.'}</p>
            </div>
            <div class="detail-block">
              <h4>Katman B · Değer / Geride Kalmışlık (${d.layer_b_score ?? '—'} puan)</h4>
              <p>${d.layer_b_notes || 'Veri yetersiz.'}</p>
            </div>
            <div class="detail-block detail-risk">
              <h4>Risk Filtresi</h4>
              <span class="chip ${tier === 'out' ? 'out' : 'watch'}">${d.risk_passed ? 'Geçti' : 'Elendi'}</span>
              <div class="risk-line">${d.risk_status || ''}</div>
            </div>
          </div>
        </td>
      </tr>`;
    }).join('');

    tbody.querySelectorAll('tr.row').forEach(tr => {
      tr.addEventListener('click', () => {
        const t = tr.dataset.ticker;
        expandedTicker = expandedTicker === t ? null : t;
        render();
      });
    });
  }

  searchInput.addEventListener('input', render);
  sectorSel.addEventListener('change', render);
  riskToggle.addEventListener('change', render);

  renderStats();
  render();
</script>
</body>
</html>
"""


def write_html_dashboard(rows, run_date_str, output_path: Path):
    """Skor + getiri verisini arama/filtre/sıralama özellikli, tek dosyalık
    bağımsız bir HTML sayfasına yazar (internet bağlantısı gerekmeden açılır,
    yalnızca Google Fonts için ağ isteği yapar)."""
    data_json = json.dumps(rows, ensure_ascii=False)
    html = DASHBOARD_TEMPLATE.replace("__BIST_DATA__", data_json).replace("__RUN_DATE__", run_date_str)
    output_path.write_text(html, encoding="utf-8")
    print(f"HTML panosu yazildi: {output_path}")


def git_publish(run_date_str):
    """docs/index.html ve history.csv dosyalarini GitHub'a commit+push eder,
    boylece GitHub Pages uzerindeki pano otomatik guncellenir. Repo/remote/
    kimlik bilgisi eksikse veya push basarisiz olursa scripti durdurmadan
    uyari basar (ilk push'un elle yapilmasi gerekebilir)."""
    import subprocess
    repo_dir = Path(__file__).resolve().parent
    try:
        subprocess.run(
            ["git", "add", "docs/index.html", "history.csv"],
            cwd=repo_dir, check=True, capture_output=True, text=True,
        )
        commit = subprocess.run(
            ["git", "commit", "-m", f"Haftalik tarama guncellemesi - {run_date_str}"],
            cwd=repo_dir, capture_output=True, text=True,
        )
        if commit.returncode != 0:
            if "nothing to commit" in commit.stdout:
                print("  Git: degisiklik yok, commit atlandi.")
            else:
                print(f"  [uyari] git commit basarisiz: {commit.stdout}{commit.stderr}")
            return
        push = subprocess.run(
            ["git", "push", "origin", "main"],
            cwd=repo_dir, capture_output=True, text=True,
        )
        if push.returncode != 0:
            print(f"  [uyari] git push basarisiz (ilk push'u elle yapman gerekebilir): {push.stderr.strip()}")
        else:
            print("  GitHub'a push edildi - Pages birkac dakika icinde guncellenir.")
    except Exception as e:
        print(f"  [uyari] Otomatik git publish basarisiz: {e}")


# =============================================================================
# ANA AKIŞ
# =============================================================================

def main():
    run_date = date.today().isoformat()
    print(f"=== BIST100 Model Portfoy Taramasi — {run_date} ===\n")

    all_companies = fetch_bulk_companies()

    if AUTO_TOP_N_BY_MARKET_CAP:
        candidates = sorted(
            all_companies.values(),
            key=lambda c: c.get("market_cap_mn_try") or 0,
            reverse=True,
        )[:TOP_N]
        tickers = [c["ticker"] for c in candidates]
    else:
        tickers = [t for t in BIST100_TICKERS if t in all_companies]
        missing = [t for t in BIST100_TICKERS if t not in all_companies]
        if missing:
            print(f"[uyari] Su kodlar bulunamadi (kod degismis/kapanmis olabilir): {missing}")

    print(f"\nToplam {len(tickers)} hisse islenecek.\n")

    sector_pb_avg = compute_sector_pb_averages(all_companies)
    history = load_history()

    rows = []
    for i, ticker in enumerate(tickers, 1):
        company = all_companies[ticker]
        print(f"[{i}/{len(tickers)}] {ticker} isleniyor...")
        try:
            detail = fetch_company_detail(ticker)
        except Exception as e:
            print(f"  [hata] {ticker} detayi alinamadi, atlaniyor: {e}")
            continue
        time.sleep(REQUEST_DELAY_SEC)

        fin = analyze_financials(detail, company.get("fin_group"))
        layer_a_score, layer_a_notes = score_layer_a(fin)

        hist_for_ticker = history.get(ticker, [])
        own_pe_ratio = compute_own_pe_ratio(hist_for_ticker, company.get("pe"),
                                             min_points=MIN_WEEKS_FOR_OWN_PE_HISTORY)
        relative_return = compute_relative_return(hist_for_ticker,
                                                    weeks_back=MIN_WEEKS_FOR_RELATIVE_RETURN)
        layer_b_score, layer_b_notes = score_layer_b(
            company, sector_pb_avg.get(company.get("sector_code")), own_pe_ratio, relative_return
        )

        risk_passed, risk_status = apply_risk_filter(company, fin)

        total_score = None
        if layer_a_score is not None:
            total_score = layer_a_score + layer_b_score

        yahoo_series = fetch_yahoo_price_history(ticker)
        time.sleep(REQUEST_DELAY_SEC)
        period_returns = compute_period_returns(yahoo_series)

        rows.append({
            "ticker": ticker,
            "name": company.get("name"),
            "sector": company.get("sector"),
            "pe": company.get("pe"),
            "pb": company.get("pb"),
            "roe": company.get("roe"),
            "layer_a_score": layer_a_score,
            "layer_a_notes": layer_a_notes or ("Banka/finans - N/A" if fin["is_financial_sector"] else "Veri yetersiz"),
            "layer_b_score": layer_b_score,
            "layer_b_notes": layer_b_notes or "Veri yetersiz",
            "total_score": total_score,
            "risk_passed": risk_passed,
            "risk_status": risk_status,
            "ret_1d": period_returns["ret_1d"],
            "ret_1w": period_returns["ret_1w"],
            "ret_1m": period_returns["ret_1m"],
            "ret_1y": period_returns["ret_1y"],
        })

    # Gecmise bu haftanin verisini ekle (BIR SONRAKI calistirmada kullanilacak)
    snapshot = {t: all_companies[t] for t in tickers}
    append_history(snapshot, run_date)

    # Excel raporu yaz
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"BIST100_Model_Portfoy_{run_date.replace('-', '')}.xlsx"
    write_excel_report(rows, run_date, out_path)

    # HTML panosu yaz (arama/filtre/siralama ozellikli, tarayicida acilan tek dosya)
    html_path = OUTPUT_DIR / f"BIST100_Model_Portfoy_{run_date.replace('-', '')}.html"
    write_html_dashboard(rows, run_date, html_path)

    # GitHub Pages icin docs/index.html'i de guncelle (sabit URL, uzaktan erisim)
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    docs_index = DOCS_DIR / "index.html"
    write_html_dashboard(rows, run_date, docs_index)

    print("\n=== Tamamlandi ===")
    print(f"Excel raporu: {out_path}")
    print(f"HTML panosu: {html_path}")
    print(f"Gecmis veri: {HISTORY_FILE} ({sum(len(v) for v in load_history().values())} toplam kayit)")

    if GIT_AUTO_PUBLISH:
        print("\nGitHub'a yayinlaniyor...")
        git_publish(run_date)


if __name__ == "__main__":
    main()
