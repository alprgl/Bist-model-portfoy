#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TEFAS FON MODEL PORTFÖY ARACI
==============================
Türkiye'de TEFAS'ta (Takasbank Fon Bilgilendirme Platformu) işlem gören tüm
yatırım fonlarını tarar, her fonu KENDİ TÜRÜNDEKİ (şemsiye fon türü) emsalleriyle
kıyaslayarak iki kriterde puanlar:

  Katman A - Getiri (%45): Günlük/haftalık/~1 aylık/3 aylık/6 aylık/1 yıllık
             getiriler ile Sharpe-benzeri risk-ayarlı bir oranın (getiri/
             volatilite), kendi türündeki fonlara göre ortalama yüzdelik dilimi.
  Katman B - Para Akışı (%20): Fonun tedavüldeki pay sayısındaki günlük değişimi
             (× o günün fiyatı) toplanarak net nakit girişi/çıkışı tahmin
             edilir, portföy büyüklüğüne oranlanır; ayrıca bu oranın fonun
             KENDİ geçmiş ortalamasına göre z-skoru hesaplanır ("normalden
             fazla/az para mı giriyor" sinyali). İkisinin ortalama percentile'ı.
  Katman C - Risk (%35): Düşük volatilite + küçük maks. düşüş + düşük yatırımcı
             yoğunlaşmasının (az sayıda büyük yatırımcıya bağımlılık) ortalama
             yüzdelik dilimi.

Toplam Skor = ağırlıklı_ortalama(Katman A×0.45, Katman B×0.20, Katman C×0.35),
sonra kendi türü içinde tekrar yüzdelik dilime çevrilir (dağılımı düz tutmak
için). Ağırlıklar eşit değil: risk-ayarlı getirinin akış sinyalinden daha
kalıcı/öngörülebilir olduğu literatür bulgusuna dayanır (Morningstar MRAR,
Refinitiv Lipper Leaders metodolojileri ve fon performansı süreklilik
literatürü).

VERİ KAYNAĞI
------------
TEFAS'ın herkese açık, anahtarsız dahili API'si:
  https://www.tefas.gov.tr/api/funds/fonGnlBlgSiraliGetir
Resmi/dökümente edilmiş bir API değildir (TEFAS'ın kendi web sitesinin
kullandığı uç nokta), bu yüzden TEFAS tarafında değişebilir.

KULLANIM
--------
    python3 fon_model_portfoy.py
"""

import calendar
import csv
import json
import time
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path

# =============================================================================
# AYARLAR
# =============================================================================

API_BASE = "https://www.tefas.gov.tr/api/funds"
FON_GENEL_URL = f"{API_BASE}/fonGnlBlgSiraliGetir"
FON_TUR_URL = f"{API_BASE}/fonTurGetir"

OUTPUT_DIR = Path("./ciktilar")
FON_SKOR_GECMIS_FILE = Path("./fon_skor_gecmis.csv")  # skor degisimini takip etmek icin
FON_AKIS_GECMIS_FILE = Path("./fon_akis_gecmis.csv")  # akis z-skoru icin fonun kendi gecmisini biriktirir
DOCS_DIR = Path(__file__).resolve().parent / "docs"   # GitHub Pages buradan yayinlanir
GIT_AUTO_PUBLISH = True    # True ise her calistirmada docs/fon.html otomatik commit+push edilir

LOOKBACK_DAYS = 30          # getiri/akis hesabi icin kac takvim gunu geriye gidilecek
RISK_MIN_PORTFOY_BUYUKLUK = 10_000_000.0   # TL - bunun altindaki fonlar ELENIR (kucuk/likit degil)
RISK_MIN_KISI_SAYISI = 10                  # bunun altindaki fonlar ELENIR (halka acik degil / ozel fon)

# Katman agirliklari: getiri kalicilik literaturunde risk-ayarli getirinin ham
# getiriden daha tutarli oldugu, akis sinyalinin ise kisa vadede daha gurultulu
# oldugu bulgularina dayanir (bkz. Morningstar MRAR, Lipper Leaders, "smart money"
# literaturu). Toplami 1.0 olmasi sart degil - eksik katmanlarda otomatik
# yeniden normalize edilir (bkz. main() icindeki agirlikli ortalama).
KATMAN_AGIRLIKLARI = {"katman_a_getiri": 0.45, "katman_b_akis": 0.20, "katman_c_risk": 0.35}

SHARPE_VOLATILITE_TABAN = 0.01   # sharpe-benzeri oranda sifira bolmeyi onlemek icin volatilite tabani (%)
AKIS_Z_MIN_GOZLEM = 10            # akis z-skoru icin fonun kendi gecmisinde en az kac gun veri gerekli

# TEFAS API tek istekte 1 aydan uzun tarih araligina izin vermiyor, bu yuzden uzun
# vadeli getiriler icin (1/3/6/12 ay) o tarihlerin etrafinda DAR birer "referans fiyat"
# penceresi ayrica cekilir (tum fonlar icin tek istek, hafif).
#
# ONEMLI: TEFAS'in kendi sitesi (tefas.gov.tr) "Getiri Bilgisi" panelindeki 1/3/6/12
# aylik degerleri SABIT GUN SAYISI (30/90/180/365) ile degil, TAKVIM AYI ile hesaplar
# (orn. "1 Ay" = bugunden tam 1 takvim ayi once, ayni gun) VE hedef tarih islem gunu
# degilse (haftasonu/tatil) EN YAKIN SONRAKI islem gununu kullanir (geriye degil,
# ILERIYE yuvarlar). Bu, TEFAS'in kendi web sitesindeki fonFiyatBilgiGetir uc noktasi
# (periyod=1/3/6/12) network istekleri incelenerek DOGRULANMISTIR - MTK ornegi icin
# 4 donemin de (1a: 31 Tem, 3a: 1 Haz [31 May Pazar->ileri], 6a: 2 Mar [28 Sub
# Cumartesi->ileri], 1y: 1 Eyl 2025 [31 Agu 2025 Pazar->ileri]) referans tarihi
# birebir eslesmistir. Asagidaki UZUN_VADE_DONEMLERI ve n_ay_once()/fetch_referans_fiyatlar()
# bu kurali uyguluyor.
UZUN_VADE_DONEMLERI = [("1a", 1), ("3a", 3), ("6a", 6), ("1y", 12)]  # (etiket, kac takvim ayi once)
REFERANS_PENCERE_GUN = 6   # hedef tarihten ILERIYE dogru kac gunluk pencereye bakilacak (tatil/haftasonu icin pay)


def n_ay_once(tarih: date, ay_sayisi: int) -> date:
    """tarih'ten ay_sayisi kadar takvim ayi once, AYNI GUNU doner. Hedef ayda o gun
    yoksa (orn. 31 Agustos - 6 ay = 31 Subat yok) ayin SON gunune sabitlenir (28/29
    Subat gibi). TEFAS'in kendi "1/3/6/12 Aylik" getiri hesabiyla ayni mantik."""
    toplam_ay = tarih.year * 12 + (tarih.month - 1) - ay_sayisi
    yil, ay = divmod(toplam_ay, 12)
    ay += 1
    son_gun = calendar.monthrange(yil, ay)[1]
    return date(yil, ay, min(tarih.day, son_gun))

REQUEST_TIMEOUT_SEC = 30
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/json",
    "Origin": "https://www.tefas.gov.tr",
    "Referer": "https://www.tefas.gov.tr/TarihselVeriler.aspx",
}


# =============================================================================
# AĞ / VERİ ÇEKME
# =============================================================================

def tefas_post(url: str, payload: dict, retries: int = 5):
    data = json.dumps(payload).encode("utf-8")
    last_err = None
    for attempt in range(1, retries + 1):
        req = urllib.request.Request(url, data=data, headers=REQUEST_HEADERS, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT_SEC) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            last_err = e
            if e.code == 429:
                wait = 3.0 * attempt
                print(f"  [uyari] 429 Too Many Requests, {wait:.0f}sn bekleniyor (deneme {attempt}/{retries})...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"{url} alinamadi (429 tekrar denemeleri tukendi): {last_err}")


def fetch_fon_turleri():
    """Semsiye fon turlerini {sfonTuru: aciklama} olarak doner."""
    payload = tefas_post(FON_TUR_URL, {"fonTipi": "YAT", "flag": 1})
    return {t["sfonTuru"]: t["sfonTurAciklama"] for t in payload.get("resultList", [])}


TUR_ANAHTAR_KELIMELER = [
    # (fonUnvan icinde aranacak anahtar kelime, sfonTuru kodu) - ilk eslesen kazanir,
    # bu yuzden daha spesifik olanlar basta.
    ("KIYMETLİ MADEN", 105), ("ALTIN", 105),
    ("PARA PİYASASI", 107),
    ("KATILIM", 114),
    ("GARANTİLİ", 103), ("KORUMA AMAÇLI", 103),
    ("FON SEPETİ", 102),
    ("SERBEST", 108),
    ("HİSSE SENEDİ", 104), ("HİSSE SENEDI", 104),
    ("BORÇLANMA", 100), ("TAHVİL", 100), ("BONO", 100),
    ("KARMA", 110),
    ("DEĞİŞKEN", 101),
]


def guess_fon_turu(fon_unvan: str):
    """Fon unvanindaki anahtar kelimelerden semsiye fon turunu tahmin eder
    (TEFAS'in tur-bazli sorgusu rate-limit'e takildigi icin bu pratik yaklasik
    kullanilir). Eslesme yoksa None doner."""
    if not fon_unvan:
        return None
    unvan_buyuk = fon_unvan.upper()
    for anahtar, sfon_tur in TUR_ANAHTAR_KELIMELER:
        if anahtar in unvan_buyuk:
            return sfon_tur
    return None


def fetch_tum_fonlar_zaman_serisi(bas_tarih: date, bit_tarih: date):
    """Tum fonlarin bas_tarih-bit_tarih arasindaki GUNLUK gecmisini TEK istekte ceker.
    Doner: {fonKodu: [{'tarih':date, 'fiyat':float, 'tedPaySayisi':float,
                        'kisiSayisi':int, 'portfoyBuyukluk':float, 'fonUnvan':str}, ...]}
    (tarihe gore ARTAN sirali)."""
    payload = {
        "fonTipi": "YAT", "fonKodu": None, "aramaMetni": None, "fonTurKod": None,
        "fonGrubu": None, "sfonTurKod": None,
        "basTarih": bas_tarih.strftime("%Y%m%d"), "bitTarih": bit_tarih.strftime("%Y%m%d"),
        "basSira": 1, "bitSira": 60000, "fonTurAciklama": None, "dil": "TR", "kurucuKod": None,
    }
    payload_json = tefas_post(FON_GENEL_URL, payload)
    rows = payload_json.get("resultList") or []

    by_fon = defaultdict(list)
    for r in rows:
        if r.get("fiyat") is None or r.get("tedPaySayisi") is None:
            continue
        by_fon[r["fonKodu"]].append({
            "tarih": datetime.strptime(r["tarih"], "%Y-%m-%d").date(),
            "fiyat": r["fiyat"],
            "tedPaySayisi": r["tedPaySayisi"],
            "kisiSayisi": r.get("kisiSayisi"),
            "portfoyBuyukluk": r.get("portfoyBuyukluk"),
            "fonUnvan": r.get("fonUnvan"),
        })
    for fon_kodu in by_fon:
        by_fon[fon_kodu].sort(key=lambda x: x["tarih"])
    return by_fon, payload_json.get("toplamSayi")


def fetch_referans_fiyatlar(hedef_tarih: date, pencere_gun: int = REFERANS_PENCERE_GUN):
    """hedef_tarih itibariyle (TEFAS'in kendi kuraliyla BIREBIR ayni: hedef_tarih islem
    gunu degilse EN YAKIN SONRAKI islem gunu) tum fonlarin referans fiyatini tek istekte
    ceker. Pencere ILERIYE dogru acilir (TEFAS 1 aydan uzun araligi reddettigi icin dar
    tutulur). Doner: {fonKodu: fiyat}."""
    bit_tarih = hedef_tarih + timedelta(days=pencere_gun)
    payload = {
        "fonTipi": "YAT", "fonKodu": None, "aramaMetni": None, "fonTurKod": None,
        "fonGrubu": None, "sfonTurKod": None,
        "basTarih": hedef_tarih.strftime("%Y%m%d"), "bitTarih": bit_tarih.strftime("%Y%m%d"),
        "basSira": 1, "bitSira": 60000, "fonTurAciklama": None, "dil": "TR", "kurucuKod": None,
    }
    payload_json = tefas_post(FON_GENEL_URL, payload)
    rows = payload_json.get("resultList") or []

    en_yakin = {}
    for r in rows:
        if r.get("fiyat") is None:
            continue
        t = datetime.strptime(r["tarih"], "%Y-%m-%d").date()
        mevcut = en_yakin.get(r["fonKodu"])
        if mevcut is None or t < mevcut[0]:   # EN ERKEN (hedef_tarih'e en yakin ILERI) tarihi tut
            en_yakin[r["fonKodu"]] = (t, r["fiyat"])
    return {kod: fiyat for kod, (_, fiyat) in en_yakin.items()}


# =============================================================================
# METRİK HESAPLAMA
# =============================================================================

def _pencere_metrikleri(series, gun_sayisi):
    """series icinde, en son noktadan gun_sayisi takvim gunu geriye giden en yakin
    noktayi bulup, o noktadan bugune: getiri % ve net akis (TL + oran %) hesaplar.
    Yetersiz veri varsa (None, None, None) doner."""
    son = series[-1]
    hedef_tarih = son["tarih"] - timedelta(days=gun_sayisi)

    nokta, nokta_idx = None, None
    for i, pt in enumerate(series):
        if pt["tarih"] <= hedef_tarih:
            nokta, nokta_idx = pt, i
        else:
            break

    if nokta is None or nokta_idx == len(series) - 1 or not nokta.get("fiyat"):
        return None, None, None

    getiri_pct = (son["fiyat"] - nokta["fiyat"]) / nokta["fiyat"] * 100.0

    net_akis_tl = 0.0
    for i in range(nokta_idx + 1, len(series)):
        onceki, simdiki = series[i - 1], series[i]
        pay_degisim = simdiki["tedPaySayisi"] - onceki["tedPaySayisi"]
        net_akis_tl += pay_degisim * simdiki["fiyat"]

    akis_oran_pct = None
    if nokta.get("portfoyBuyukluk") and nokta["portfoyBuyukluk"] > 0:
        akis_oran_pct = net_akis_tl / nokta["portfoyBuyukluk"] * 100.0

    return getiri_pct, net_akis_tl, akis_oran_pct


def _pencere_risk_metrikleri(series, gun_sayisi):
    """series icinde son gun_sayisi takvim gunundeki GUNLUK fiyat degisimlerinden
    volatilite (gunluk getirilerin std sapmasi, %) ve maksimum dusus (en yuksek
    noktadan en buyuk % geri cekilme, negatif deger) hesaplar. Doner: (volatilite,
    max_dusus) - yetersiz veri varsa (None, None)."""
    son_tarih = series[-1]["tarih"]
    hedef_tarih = son_tarih - timedelta(days=gun_sayisi)
    pencere = [pt for pt in series if pt["tarih"] > hedef_tarih and pt.get("fiyat")]
    if len(pencere) < 3:
        return None, None

    fiyatlar = [pt["fiyat"] for pt in pencere]
    getiriler = [(fiyatlar[i] - fiyatlar[i - 1]) / fiyatlar[i - 1]
                 for i in range(1, len(fiyatlar)) if fiyatlar[i - 1]]
    volatilite = None
    if len(getiriler) >= 2:
        ort = sum(getiriler) / len(getiriler)
        varyans = sum((g - ort) ** 2 for g in getiriler) / (len(getiriler) - 1)
        volatilite = (varyans ** 0.5) * 100.0

    tepe = fiyatlar[0]
    max_dusus = 0.0
    for f in fiyatlar:
        if f > tepe:
            tepe = f
        dusus = (f - tepe) / tepe * 100.0
        if dusus < max_dusus:
            max_dusus = dusus

    return volatilite, max_dusus


def compute_fund_metrics(series):
    """series: tarihe gore artan sirali gunluk kayitlar (bkz. fetch_tum_fonlar_zaman_serisi).
    Gunluk (1g), haftalik (1h) ve ~LOOKBACK_DAYS gunluk (1a) pencereler icin ayri ayri
    getiri/akis hesaplar - hepsi zaten cekilmis tek seriden turetilir, ekstra istek gerekmez.
    Doner: dict veya None (yetersiz veri)."""
    if len(series) < 2:
        return None

    ilk, son = series[0], series[-1]
    if not ilk["fiyat"] or ilk["fiyat"] <= 0:
        return None

    getiri_1a, net_akis_1a, akis_oran_1a = _pencere_metrikleri(series, LOOKBACK_DAYS)
    getiri_1h, net_akis_1h, akis_oran_1h = _pencere_metrikleri(series, 7)
    getiri_1g, net_akis_1g, akis_oran_1g = _pencere_metrikleri(series, 1)

    # 1 aylik pencere, seride en az iki nokta oldugu surece serinin ilk noktasina
    # kadar geriye gitmeyi garanti eder (asagidaki fallback ile).
    if getiri_1a is None:
        getiri_1a = (son["fiyat"] - ilk["fiyat"]) / ilk["fiyat"] * 100.0
        net_akis_1a = 0.0
        for i in range(1, len(series)):
            onceki, simdiki = series[i - 1], series[i]
            net_akis_1a += (simdiki["tedPaySayisi"] - onceki["tedPaySayisi"]) * simdiki["fiyat"]
        akis_oran_1a = (net_akis_1a / ilk["portfoyBuyukluk"] * 100.0
                         if ilk.get("portfoyBuyukluk") else None)

    kisi_degisim = None
    if ilk.get("kisiSayisi") is not None and son.get("kisiSayisi") is not None:
        kisi_degisim = son["kisiSayisi"] - ilk["kisiSayisi"]

    volatilite_1h, max_dusus_1h = _pencere_risk_metrikleri(series, 7)
    volatilite_1a, max_dusus_1a = _pencere_risk_metrikleri(series, LOOKBACK_DAYS)

    # Sharpe-benzeri risk-ayarli oran: ~1 aylik getiri / ~1 aylik volatilite. Risksiz
    # faiz orani dusulmedigi ve gunluk getiri std'si yillandirilmadigi icin GERCEK bir
    # Sharpe orani degildir - sadece AYNI TUR icindeki fonlari "getiri basina risk"
    # acisindan siralamak icin kullanilir (mutlak degeri degil, goreceli sirasi onemli).
    sharpe_1a = None
    if volatilite_1a is not None:
        sharpe_1a = getiri_1a / max(volatilite_1a, SHARPE_VOLATILITE_TABAN)

    yogunlasma_tl = None  # ortalama yatirimci payi (TL) - yuksekse birkac buyuk yatirimciya bagimli
    if son.get("kisiSayisi") and son["kisiSayisi"] > 0 and son.get("portfoyBuyukluk"):
        yogunlasma_tl = son["portfoyBuyukluk"] / son["kisiSayisi"]

    return {
        "fonUnvan": son.get("fonUnvan"),
        "guncel_fiyat": son["fiyat"],
        "guncel_portfoy_buyuklugu": son.get("portfoyBuyukluk"),
        "guncel_kisi_sayisi": son.get("kisiSayisi"),
        "getiri_pct": getiri_1a, "net_akis_tl": net_akis_1a, "akis_oran_pct": akis_oran_1a,
        "getiri_1g": getiri_1g, "net_akis_1g": net_akis_1g, "akis_oran_1g": akis_oran_1g,
        "getiri_1h": getiri_1h, "net_akis_1h": net_akis_1h, "akis_oran_1h": akis_oran_1h,
        "volatilite_1h": volatilite_1h, "max_dusus_1h": max_dusus_1h,
        "volatilite_1a": volatilite_1a, "max_dusus_1a": max_dusus_1a,
        "sharpe_1a": sharpe_1a,
        "yogunlasma_tl": yogunlasma_tl,
        "kisi_degisim": kisi_degisim,
        "veri_gun_sayisi": len(series),
        "ilk_tarih": ilk["tarih"].isoformat(),
        "son_tarih": son["tarih"].isoformat(),
    }


def apply_risk_filter(metrics):
    reasons = []
    pb = metrics.get("guncel_portfoy_buyuklugu")
    ks = metrics.get("guncel_kisi_sayisi")
    if pb is not None and pb < RISK_MIN_PORTFOY_BUYUKLUK:
        reasons.append(f"Portföy büyüklüğü {pb:,.0f} TL < {RISK_MIN_PORTFOY_BUYUKLUK:,.0f} TL")
    if ks is not None and ks < RISK_MIN_KISI_SAYISI:
        reasons.append(f"Yatırımcı sayısı {ks} < {RISK_MIN_KISI_SAYISI}")
    passed = len(reasons) == 0
    return passed, "; ".join(reasons) if reasons else "GEÇTİ"


def percentile_rank(value, all_values_sorted):
    """value'nun all_values_sorted (artan sirali) icindeki yuzdelik dilimini (0-100) doner."""
    if value is None or not all_values_sorted:
        return None
    n = len(all_values_sorted)
    # kacinin value'dan kucuk oldugunu say (basit ama 2000 fon icin yeterince hizli)
    kucuk_sayisi = sum(1 for v in all_values_sorted if v < value)
    esit_sayisi = sum(1 for v in all_values_sorted if v == value)
    return (kucuk_sayisi + 0.5 * esit_sayisi) / n * 100.0


# =============================================================================
# VERİ KALİTESİ KONTROLÜ
# =============================================================================
#
# Her calistirmada otomatik calisan bir "mantik kontrolu" katmani. TEFAS'in
# kendi sitesiyle tek tek manuel karsilastirma yapmanin yerini almaz (o hala
# en guvenilir dogrulama yontemidir), ama HER calistirmada, HER fon icin,
# bariz/imkansiz degerleri (negatif buyukluk, asiri hareket, bayat veri,
# tekrar eden fon kodu vb.) otomatik yakalar - boylece hatali bir veri sessizce
# rapora sizmez, gorunur bir uyari olarak isaretlenir.

VERI_KALITE_ESIKLERI = {
    "getiri_1g_mutlak_maks": 40.0,    # % - tek gunde bu kadar buyuk hareket suphelidir
    "getiri_pct_mutlak_maks": 150.0,  # % - ~1 ayda bu kadar buyuk hareket suphelidir
    "akis_oran_mutlak_maks": 500.0,   # % - ~1 ayda buyuklugunun bu katini akis olarak degistirmek suphelidir
    "veri_bayat_gun": 5,              # son veri noktasi run_date'den bu kadar eskiyse "guncel degil" sayilir
}

# Her uyari turunun HANGI sutunda gosterilecegini belirler - boylece "Veri
# Uyarisi" metnini okumak yerine, sorunlu HUCRENIN kendisi Excel'de/panoda
# vurgulanabilir. (excel_sutun, dashboard_alan) - write_excel_report ve
# DASHBOARD_TEMPLATE bu listeyi referans alir, JS tarafinda ayni liste elle
# senkron tutulur (bkz. DASHBOARD_TEMPLATE icindeki UYARI_ALAN_ESLEME).
UYARI_ALAN_ESLEME = [
    ("Geçersiz güncel fiyat", "Güncel Fiyat", "guncel_fiyat"),
    ("Negatif portföy büyüklüğü", "Portföy Büyüklüğü (Mn TL)", "portfoy_buyuklugu_mn"),
    ("Negatif yatırımcı sayısı", "Yatırımcı Sayısı", "kisi_sayisi"),
    ("Günlük getiri aşırı", "Günlük Getiri %", "getiri_1g"),
    ("~1 Ay getiri aşırı", "~1 Ay Getiri %", "getiri_pct"),
    ("Akış oranı aşırı", "Akış/Büyüklük % (1A)", "akis_oran_pct"),
    ("Negatif volatilite", "Volatilite % (1A)", "volatilite_1a"),
    ("Veri güncel değil", "Son Veri Tarihi", "son_tarih"),
    ("Fon kodu taramada birden fazla kez geçiyor", "Kod", "fonKodu"),
]


def validate_data_quality(ham_sonuclar, run_date: date):
    """Her fonun HAM verisini (skorlama ONCESI) basit mantik kontrollerinden
    gecirir. Skorlama modelinin isabetini degil, ALTINDAKI VERININ makul olup
    olmadigini kontrol eder. Her satirin 'veri_uyarilari' alanina bulunan
    sorunlarin listesini yazar (bos liste = sorun yok), ayrica sadece sorunlu
    fonlarin ozetini bir liste olarak doner."""
    esik = VERI_KALITE_ESIKLERI

    kod_sayaci = defaultdict(int)
    for r in ham_sonuclar:
        kod_sayaci[r["fonKodu"]] += 1

    ozet = []
    for r in ham_sonuclar:
        uyarilar = []

        if r.get("guncel_fiyat") is None or r["guncel_fiyat"] <= 0:
            uyarilar.append("Geçersiz güncel fiyat")

        pb = r.get("guncel_portfoy_buyuklugu")
        if pb is not None and pb < 0:
            uyarilar.append("Negatif portföy büyüklüğü")

        ks = r.get("guncel_kisi_sayisi")
        if ks is not None and ks < 0:
            uyarilar.append("Negatif yatırımcı sayısı")

        g1g = r.get("getiri_1g")
        if g1g is not None and abs(g1g) > esik["getiri_1g_mutlak_maks"]:
            uyarilar.append(f"Günlük getiri aşırı: {g1g:+.1f}%")

        g1a = r.get("getiri_pct")
        if g1a is not None and abs(g1a) > esik["getiri_pct_mutlak_maks"]:
            uyarilar.append(f"~1 Ay getiri aşırı: {g1a:+.1f}%")

        akis = r.get("akis_oran_pct")
        if akis is not None and abs(akis) > esik["akis_oran_mutlak_maks"]:
            uyarilar.append(f"Akış oranı aşırı: {akis:+.1f}%")

        vol = r.get("volatilite_1a")
        if vol is not None and vol < 0:
            uyarilar.append("Negatif volatilite (hesaplama hatası)")

        son_tarih_str = r.get("son_tarih")
        if son_tarih_str:
            son_tarih = date.fromisoformat(son_tarih_str)
            if (run_date - son_tarih).days > esik["veri_bayat_gun"]:
                uyarilar.append(f"Veri güncel değil (son veri: {son_tarih_str})")

        if kod_sayaci[r["fonKodu"]] > 1:
            uyarilar.append("Fon kodu taramada birden fazla kez geçiyor")

        r["veri_uyarilari"] = uyarilar
        if uyarilar:
            ozet.append({"fonKodu": r["fonKodu"], "fonUnvan": r.get("fonUnvan"), "uyarilar": uyarilar})

    return ozet


# =============================================================================
# SKOR GEÇMİŞİ (günlük/haftalık skor değişimini takip etmek için)
# =============================================================================

def load_fon_skor_gecmis():
    """fon_skor_gecmis.csv'yi {fonKodu: [{'tarih': date, 'toplam_skor': float}, ...]}
    seklinde okur (tarihe gore artan sirali). Dosya yoksa bos dict doner."""
    if not FON_SKOR_GECMIS_FILE.exists():
        return {}
    gecmis = defaultdict(list)
    with open(FON_SKOR_GECMIS_FILE, "r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            if row.get("toplam_skor") in (None, "", "None"):
                continue
            gecmis[row["fonKodu"]].append({
                "tarih": datetime.strptime(row["tarih"], "%Y-%m-%d").date(),
                "toplam_skor": float(row["toplam_skor"]),
            })
    for fon_kodu in gecmis:
        gecmis[fon_kodu].sort(key=lambda x: x["tarih"])
    return gecmis


def append_fon_skor_gecmis(ham_sonuclar, run_date: date):
    """Bu calistirmanin skorlarini fon_skor_gecmis.csv'ye ekler. Ayni gun icin
    zaten kayit varsa tekrar eklemez (ayni gun birden fazla calistirma icin)."""
    file_exists = FON_SKOR_GECMIS_FILE.exists()
    if file_exists:
        with open(FON_SKOR_GECMIS_FILE, "r", encoding="utf-8", newline="") as f:
            mevcut_tarihler = {row["tarih"] for row in csv.DictReader(f)}
        if run_date.isoformat() in mevcut_tarihler:
            print(f"  fon_skor_gecmis.csv: {run_date} tarihli kayit zaten var, tekrar eklenmedi.")
            return
    with open(FON_SKOR_GECMIS_FILE, "a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["tarih", "fonKodu", "toplam_skor"])
        for r in ham_sonuclar:
            writer.writerow([run_date.isoformat(), r["fonKodu"],
                              r["toplam_skor"] if r["toplam_skor"] is not None else ""])
    print(f"fon_skor_gecmis.csv guncellendi (+{len(ham_sonuclar)} satir, tarih={run_date}).")


def compute_skor_degisim(gecmis_for_fon, bugunku_skor, run_date: date, gun_sayisi: int):
    """gecmis_for_fon icinde run_date'den gun_sayisi takvim gunu once (veya en yakin
    onceki) kayitli skoru bulup, bugunku_skor ile arasindaki farki doner. Yetersiz
    gecmis varsa None doner."""
    if bugunku_skor is None or not gecmis_for_fon:
        return None
    hedef_tarih = run_date - timedelta(days=gun_sayisi)
    nokta = None
    for kayit in gecmis_for_fon:
        if kayit["tarih"] <= hedef_tarih:
            nokta = kayit
        else:
            break
    if nokta is None:
        return None
    return bugunku_skor - nokta["toplam_skor"]


# =============================================================================
# AKIŞ GEÇMİŞİ (fonun akışını KENDİ geçmişine göre z-skora çevirmek için)
# =============================================================================
#
# Literatürdeki "smart money" bulgusu, ham para girişinden çok BEKLENENDEN
# FAZLA/AZ girişin sinyal taşıdığını gösteriyor. Burada "beklenen akış", fonun
# kendi geçmiş ~1 aylık akış/büyüklük oranlarının ortalaması olarak alınıyor;
# bugünkü oran bu ortalamaya göre kaç standart sapma uzakta (z-skoru) hesaplanır.
# Böylece yapısal olarak hep para toplayan/kaybeden bir fon değil, NORMALİNDEN
# SAPAN fonlar öne çıkar.

def load_fon_akis_gecmis():
    """fon_akis_gecmis.csv'yi {fonKodu: [{'tarih': date, 'akis_oran_pct': float}, ...]}
    seklinde okur (tarihe gore artan sirali). Dosya yoksa bos dict doner."""
    if not FON_AKIS_GECMIS_FILE.exists():
        return {}
    gecmis = defaultdict(list)
    with open(FON_AKIS_GECMIS_FILE, "r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            if row.get("akis_oran_pct") in (None, "", "None"):
                continue
            gecmis[row["fonKodu"]].append({
                "tarih": datetime.strptime(row["tarih"], "%Y-%m-%d").date(),
                "akis_oran_pct": float(row["akis_oran_pct"]),
            })
    for fon_kodu in gecmis:
        gecmis[fon_kodu].sort(key=lambda x: x["tarih"])
    return gecmis


def append_fon_akis_gecmis(ham_sonuclar, run_date: date):
    """Bu calistirmanin ~1 aylik akis/buyukluk oranlarini fon_akis_gecmis.csv'ye
    ekler (akis z-skoru icin gerekli gecmis birikimi). Ayni gun icin zaten kayit
    varsa tekrar eklemez."""
    file_exists = FON_AKIS_GECMIS_FILE.exists()
    if file_exists:
        with open(FON_AKIS_GECMIS_FILE, "r", encoding="utf-8", newline="") as f:
            mevcut_tarihler = {row["tarih"] for row in csv.DictReader(f)}
        if run_date.isoformat() in mevcut_tarihler:
            print(f"  fon_akis_gecmis.csv: {run_date} tarihli kayit zaten var, tekrar eklenmedi.")
            return
    with open(FON_AKIS_GECMIS_FILE, "a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["tarih", "fonKodu", "akis_oran_pct"])
        for r in ham_sonuclar:
            akis = r.get("akis_oran_pct")
            writer.writerow([run_date.isoformat(), r["fonKodu"], akis if akis is not None else ""])
    print(f"fon_akis_gecmis.csv guncellendi (+{len(ham_sonuclar)} satir, tarih={run_date}).")


def compute_akis_z(gecmis_for_fon, bugunku_akis, min_gozlem: int = AKIS_Z_MIN_GOZLEM):
    """Fonun BUGUNKU ~1 aylik akis/buyukluk oranini, KENDI GECMISINDEKI (bugun
    haric) ayni oranin ortalama ve standart sapmasina gore z-skora cevirir: pozitif
    z = fona normalden fazla, negatif z = normalden az para giriyor. Gecmis
    min_gozlem'den az veya varyans sifirsa None doner (ilk ~min_gozlem calistirma
    icin bos gelir - skor_degisim alanlariyla ayni birikim mantigi)."""
    if bugunku_akis is None or len(gecmis_for_fon) < min_gozlem:
        return None
    degerler = [k["akis_oran_pct"] for k in gecmis_for_fon]
    ort = sum(degerler) / len(degerler)
    varyans = sum((d - ort) ** 2 for d in degerler) / (len(degerler) - 1)
    std = varyans ** 0.5
    if std <= 1e-9:
        return None
    return (bugunku_akis - ort) / std


# =============================================================================
# HTML PANO (bagimsiz, tek dosyalik interaktif rapor)
# =============================================================================

DASHBOARD_TEMPLATE = """<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>TEFAS Fon Model Portföy</title>

<style>
  :root {
    --bg: #F5F5F7;
    --bg-elevated: #FFFFFF;
    --bg-sunken: #F5F5F7;
    --ink: #1D1D1F;
    --ink-muted: #6E6E73;
    --ink-faint: #AEAEB2;
    --border: rgba(0, 0, 0, 0.08);
    --border-strong: rgba(0, 0, 0, 0.14);
    --accent: #248A3D;
    --accent-fill: #34C759;
    --accent-ink: #FFFFFF;
    --accent-soft: rgba(52, 199, 89, 0.13);
    --positive: #248A3D;
    --positive-bg: rgba(52, 199, 89, 0.13);
    --negative: #D70015;
    --negative-bg: rgba(255, 59, 48, 0.12);
    --tier-strong-bg: rgba(52, 199, 89, 0.15);
    --tier-strong-ink: #1F7A37;
    --tier-watch-bg: rgba(255, 149, 0, 0.14);
    --tier-watch-ink: #A15C00;
    --tier-weak-bg: rgba(110, 110, 115, 0.09);
    --tier-weak-ink: #6E6E73;
    --tier-out-bg: rgba(255, 59, 48, 0.11);
    --tier-out-ink: #D70015;
    --star: #FF9500;
    --header-glass: rgba(245, 245, 247, 0.72);
    --shadow: 0 1px 2px rgba(0, 0, 0, 0.04), 0 12px 28px -14px rgba(0, 0, 0, 0.16);
    --font-display: -apple-system, BlinkMacSystemFont, 'SF Pro Display', 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
    --font-body: -apple-system, BlinkMacSystemFont, 'SF Pro Text', 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
    --font-mono: ui-monospace, 'SF Mono', Menlo, Consolas, monospace;
  }

  @media (prefers-color-scheme: dark) {
    :root:not([data-theme="light"]) {
      --bg: #000000;
      --bg-elevated: #1C1C1E;
      --bg-sunken: #000000;
      --ink: #F5F5F7;
      --ink-muted: #98989D;
      --ink-faint: #636366;
      --border: rgba(255, 255, 255, 0.09);
      --border-strong: rgba(255, 255, 255, 0.16);
      --accent: #30D158;
      --accent-fill: #30D158;
      --accent-ink: #00390D;
      --accent-soft: rgba(48, 209, 88, 0.16);
      --positive: #30D158;
      --positive-bg: rgba(48, 209, 88, 0.14);
      --negative: #FF453A;
      --negative-bg: rgba(255, 69, 58, 0.14);
      --tier-strong-bg: rgba(48, 209, 88, 0.18);
      --tier-strong-ink: #30D158;
      --tier-watch-bg: rgba(255, 214, 10, 0.16);
      --tier-watch-ink: #FFD60A;
      --tier-weak-bg: rgba(152, 152, 157, 0.12);
      --tier-weak-ink: #98989D;
      --tier-out-bg: rgba(255, 69, 58, 0.14);
      --tier-out-ink: #FF453A;
      --star: #FFD60A;
      --header-glass: rgba(28, 28, 30, 0.72);
      --shadow: 0 1px 2px rgba(0, 0, 0, 0.3), 0 16px 36px -16px rgba(0, 0, 0, 0.65);
    }
  }

  :root[data-theme="dark"] {
    --bg: #000000;
    --bg-elevated: #1C1C1E;
    --bg-sunken: #000000;
    --ink: #F5F5F7;
    --ink-muted: #98989D;
    --ink-faint: #636366;
    --border: rgba(255, 255, 255, 0.09);
    --border-strong: rgba(255, 255, 255, 0.16);
    --accent: #30D158;
    --accent-fill: #30D158;
    --accent-ink: #00390D;
    --accent-soft: rgba(48, 209, 88, 0.16);
    --positive: #30D158;
    --positive-bg: rgba(48, 209, 88, 0.14);
    --negative: #FF453A;
    --negative-bg: rgba(255, 69, 58, 0.14);
    --tier-strong-bg: rgba(48, 209, 88, 0.18);
    --tier-strong-ink: #30D158;
    --tier-watch-bg: rgba(255, 214, 10, 0.16);
    --tier-watch-ink: #FFD60A;
    --tier-weak-bg: rgba(152, 152, 157, 0.12);
    --tier-weak-ink: #98989D;
    --tier-out-bg: rgba(255, 69, 58, 0.14);
    --tier-out-ink: #FF453A;
    --star: #FFD60A;
    --header-glass: rgba(28, 28, 30, 0.72);
    --shadow: 0 1px 2px rgba(0, 0, 0, 0.3), 0 16px 36px -16px rgba(0, 0, 0, 0.65);
  }

  * { box-sizing: border-box; }
  html, body { height: 100%; }
  body {
    margin: 0;
    background: var(--bg);
    color: var(--ink);
    font-family: var(--font-body);
    font-size: 15px;
    line-height: 1.45;
    -webkit-font-smoothing: antialiased;
  }
  @media (prefers-reduced-motion: reduce) {
    * { animation-duration: 0.001ms !important; transition-duration: 0.001ms !important; }
  }

  .app {
    display: flex;
    flex-direction: column;
    height: 100vh;
    max-width: 1440px;
    margin: 0 auto;
    padding: 0 clamp(14px, 3vw, 32px);
  }

  .masthead {
    display: flex;
    flex-wrap: wrap;
    align-items: baseline;
    justify-content: space-between;
    gap: 8px 24px;
    padding: 26px 2px 18px;
  }
  .masthead-id { display: flex; flex-direction: column; gap: 3px; }
  .eyebrow {
    font-family: var(--font-body);
    font-size: 12.5px;
    font-weight: 590;
    letter-spacing: -0.01em;
    color: var(--accent);
  }
  h1 {
    margin: 0;
    font-family: var(--font-display);
    font-weight: 700;
    font-size: clamp(24px, 3vw, 32px);
    letter-spacing: -0.025em;
    text-wrap: balance;
  }
  .masthead-meta {
    text-align: right;
    font-size: 12.5px;
    color: var(--ink-muted);
    line-height: 1.5;
  }
  .masthead-meta strong { color: var(--ink); font-weight: 600; }

  .stats {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(130px, 1fr));
    gap: 10px;
    margin: 6px 0 16px;
  }
  .stat {
    background: var(--bg-elevated);
    border-radius: 16px;
    padding: 14px 16px;
    display: flex;
    flex-direction: column;
    gap: 4px;
    box-shadow: var(--shadow);
  }
  .stat-label {
    font-size: 11px;
    letter-spacing: -0.005em;
    color: var(--ink-muted);
    font-weight: 590;
  }
  .stat-value {
    font-family: var(--font-display);
    font-size: 22px;
    font-weight: 700;
    letter-spacing: -0.02em;
    font-variant-numeric: tabular-nums;
    color: var(--ink);
  }
  .stat-value.pos { color: var(--positive); }
  .stat-value.neg { color: var(--negative); }
  .stat-sub { font-size: 11.5px; color: var(--ink-faint); }

  .toolbar {
    display: flex;
    flex-wrap: wrap;
    align-items: center;
    gap: 10px;
    padding-bottom: 12px;
  }
  .search-field {
    position: relative;
    flex: 1 1 220px;
    min-width: 160px;
  }
  .search-field input {
    width: 100%;
    padding: 9px 14px 9px 34px;
    border-radius: 980px;
    border: 1px solid var(--border);
    background: var(--bg-elevated);
    color: var(--ink);
    font-family: var(--font-body);
    font-size: 13.5px;
    box-shadow: var(--shadow);
  }
  .search-field input:focus-visible { outline: 2px solid var(--accent); outline-offset: 1px; }
  .search-field::before {
    content: "";
    position: absolute;
    left: 13px; top: 50%;
    width: 12px; height: 12px;
    margin-top: -6px;
    border: 1.5px solid var(--ink-faint);
    border-radius: 50%;
    box-shadow: 5px 5px 0 -3px var(--ink-faint);
  }
  select {
    padding: 9px 14px;
    border-radius: 980px;
    border: 1px solid var(--border);
    background: var(--bg-elevated);
    color: var(--ink);
    font-family: var(--font-body);
    font-size: 13px;
    cursor: pointer;
    box-shadow: var(--shadow);
  }
  select:focus-visible { outline: 2px solid var(--accent); outline-offset: 1px; }
  .result-count {
    margin-left: auto;
    font-size: 12px;
    color: var(--ink-muted);
    font-variant-numeric: tabular-nums;
  }

  .switch-row {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 6px 14px 6px 6px;
    border-radius: 980px;
    background: var(--bg-elevated);
    border: 1px solid var(--border);
    box-shadow: var(--shadow);
    user-select: none;
    cursor: pointer;
    white-space: nowrap;
  }
  .switch-row-label { font-size: 13px; color: var(--ink); }
  .switch { position: relative; display: inline-block; width: 38px; height: 23px; flex-shrink: 0; }
  .switch input {
    position: absolute; inset: 0; width: 100%; height: 100%;
    margin: 0; opacity: 0; cursor: pointer;
  }
  .switch-track {
    position: absolute; inset: 0;
    background: var(--ink-faint);
    opacity: 0.32;
    border-radius: 999px;
    transition: background 0.2s ease, opacity 0.2s ease;
  }
  .switch-knob {
    position: absolute; top: 2px; left: 2px;
    width: 19px; height: 19px;
    border-radius: 50%;
    background: #FFFFFF;
    box-shadow: 0 1px 1px rgba(0,0,0,0.2), 0 2px 6px rgba(0,0,0,0.15);
    transition: transform 0.22s cubic-bezier(0.4, 0, 0.2, 1);
  }
  .switch input:checked ~ .switch-track { background: var(--accent-fill); opacity: 1; }
  .switch input:checked ~ .switch-track .switch-knob { transform: translateX(15px); }
  .switch input:focus-visible ~ .switch-track { outline: 2px solid var(--accent); outline-offset: 2px; }

  .range-field {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 4px 6px 4px 4px;
    border-radius: 980px;
    border: 1px solid var(--border);
    background: var(--bg-elevated);
    box-shadow: var(--shadow);
  }
  .range-field.active { border-color: var(--accent); }
  .range-field select {
    box-shadow: none;
    border: none;
    background: transparent;
    padding: 6px 4px;
    max-width: 160px;
  }
  .range-field input[type="number"] {
    width: 60px;
    padding: 6px 8px;
    border-radius: 8px;
    border: 1px solid var(--border-strong);
    background: var(--bg-sunken);
    color: var(--ink);
    font-family: var(--font-body);
    font-size: 13px;
    font-variant-numeric: tabular-nums;
  }
  .range-field input[type="number"]:focus-visible { outline: 2px solid var(--accent); outline-offset: 1px; }
  .range-sep { color: var(--ink-faint); font-size: 12px; }
  .range-clear {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 20px; height: 20px;
    border-radius: 50%;
    border: none;
    background: transparent;
    color: var(--ink-faint);
    font-size: 15px;
    line-height: 1;
    cursor: pointer;
    flex-shrink: 0;
  }
  .range-clear:hover { background: var(--bg-sunken); color: var(--ink); }

  .range-add {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    padding: 8px 16px;
    border-radius: 980px;
    border: none;
    background: var(--accent-fill);
    color: var(--accent-ink);
    font-family: var(--font-body);
    font-size: 12.5px;
    font-weight: 590;
    cursor: pointer;
    flex-shrink: 0;
  }
  .range-add:hover { filter: brightness(1.06); }
  .range-add:disabled { opacity: 0.35; cursor: not-allowed; filter: none; }

  .filter-chips {
    display: none;
    flex-wrap: wrap;
    align-items: center;
    gap: 8px;
    padding: 0 0 12px;
  }
  .filter-chips.visible { display: flex; }
  .filter-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 5px 6px 5px 12px;
    border-radius: 980px;
    background: var(--accent-soft);
    color: var(--accent);
    font-size: 12.5px;
    font-weight: 590;
    font-family: var(--font-body);
    font-variant-numeric: tabular-nums;
    white-space: nowrap;
  }
  .filter-chip-remove {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 18px; height: 18px;
    border-radius: 50%;
    border: none;
    background: none;
    color: var(--accent);
    font-size: 14px;
    line-height: 1;
    cursor: pointer;
    opacity: 0.75;
  }
  .filter-chip-remove:hover { opacity: 1; background: rgba(0,0,0,0.08); }
  .filter-chips-clear {
    background: none;
    border: none;
    color: var(--ink-faint);
    font-size: 12px;
    text-decoration: underline;
    cursor: pointer;
    padding: 4px;
  }
  .filter-chips-clear:hover { color: var(--ink); }

  .table-wrap {
    flex: 1;
    min-height: 0;
    overflow: auto;
    border: 1px solid var(--border);
    border-radius: 16px;
    background: var(--bg-elevated);
    box-shadow: var(--shadow);
  }
  table { border-collapse: collapse; width: 100%; min-width: 2400px; }
  thead th {
    position: sticky; top: 0; z-index: 2;
    background: var(--header-glass);
    -webkit-backdrop-filter: blur(20px) saturate(180%);
    backdrop-filter: blur(20px) saturate(180%);
    border-bottom: 1px solid var(--border);
    text-align: left;
    padding: 10px 12px;
    font-size: 11px;
    color: var(--ink-muted);
    font-weight: 590;
    letter-spacing: -0.005em;
    white-space: nowrap;
    cursor: pointer;
  }
  thead th:hover { color: var(--ink); }
  thead th.active { color: var(--accent); }
  thead th .arrow { font-size: 9px; margin-left: 3px; opacity: 0.8; }
  thead th.num, td.num { text-align: right; }
  tbody tr.row { border-bottom: 1px solid var(--border); }
  tbody tr.row:hover { background: var(--accent-soft); }
  tbody tr.row.out { opacity: 0.55; }
  td { padding: 8px 12px; font-size: 13px; white-space: nowrap; }
  td.ticker { font-weight: 590; letter-spacing: -0.005em; font-variant-numeric: tabular-nums; }
  td.name { white-space: normal; min-width: 220px; }
  td.name .sector { display: block; font-size: 11px; color: var(--ink-faint); font-weight: 400; margin-top: 1px; }
  td.num { font-variant-numeric: tabular-nums; color: var(--ink-muted); }
  .ret.pos { color: var(--positive); }
  .ret.neg { color: var(--negative); }
  .ret.zero { color: var(--ink-faint); }

  .score-cell { display: flex; align-items: center; gap: 8px; justify-content: flex-end; }
  .score-num { font-weight: 590; font-variant-numeric: tabular-nums; width: 30px; text-align: right; }
  .score-track { width: 46px; height: 6px; border-radius: 4px; background: var(--bg-sunken); overflow: hidden; flex-shrink: 0; }
  .score-fill { height: 100%; border-radius: 4px; background: var(--ink-faint); }
  .row.strong .score-fill { background: var(--tier-strong-ink); }
  .row.strong .score-num { color: var(--tier-strong-ink); }
  .row.watch .score-fill { background: var(--tier-watch-ink); }
  .row.watch .score-num { color: var(--tier-watch-ink); }
  .row.weak .score-fill { background: var(--ink-faint); }
  .row.out .score-fill { background: var(--tier-out-ink); }

  .chip {
    display: inline-block;
    padding: 2px 9px;
    border-radius: 980px;
    font-size: 10.5px;
    font-weight: 590;
    letter-spacing: -0.005em;
    white-space: nowrap;
  }
  .chip.strong { background: var(--tier-strong-bg); color: var(--tier-strong-ink); }
  .chip.watch { background: var(--tier-watch-bg); color: var(--tier-watch-ink); }
  .chip.weak { background: var(--tier-weak-bg); color: var(--tier-weak-ink); }
  .chip.out { background: var(--tier-out-bg); color: var(--tier-out-ink); }

  td.cell-warn {
    background: var(--tier-out-bg);
    box-shadow: inset 0 0 0 1px var(--tier-out-ink);
    color: var(--tier-out-ink);
  }

  td.fav-cell { padding: 8px 6px; width: 30px; }
  .fav-star {
    background: none;
    border: none;
    cursor: pointer;
    font-size: 16px;
    line-height: 1;
    padding: 4px;
    color: var(--ink-faint);
    transition: color 0.15s, transform 0.1s;
  }
  .fav-star:hover { color: var(--star); transform: scale(1.15); }
  .fav-star.active { color: var(--star); }
  .fav-star:focus-visible { outline: 2px solid var(--accent); outline-offset: 1px; border-radius: 4px; }

  .empty-state {
    padding: 48px 20px;
    text-align: center;
    color: var(--ink-muted);
    font-size: 13.5px;
  }

  footer {
    padding: 10px 2px 16px;
    font-size: 11.5px;
    color: var(--ink-faint);
    border-top: 1px solid var(--border);
    display: flex;
    flex-wrap: wrap;
    justify-content: space-between;
    gap: 6px 20px;
  }

  ::-webkit-scrollbar { width: 10px; height: 10px; }
  ::-webkit-scrollbar-track { background: transparent; }
  ::-webkit-scrollbar-thumb { background: var(--border-strong); border-radius: 6px; }

  @media (max-width: 760px) {
    .app { height: auto; min-height: 100vh; }
    .table-wrap { flex: none; }
  }
</style>
</head>
<body>
<div class="app">
  <header class="masthead">
    <div class="masthead-id">
      <span class="eyebrow">TEFAS · Fon Taraması</span>
      <h1>TEFAS</h1>
    </div>
    <div class="masthead-meta">
      Tarama tarihi: <strong id="run-date">—</strong><br>
      Kaynak: tefas.gov.tr (Takasbank)
    </div>
  </header>

  <section class="stats" id="stats"></section>

  <section class="toolbar">
    <div class="search-field">
      <input id="search" type="text" placeholder="Kod veya fon adı ara…" autocomplete="off">
    </div>
    <select id="tur-filter"></select>
    <div class="range-field" id="range-field">
      <select id="range-key"></select>
      <input id="range-min" type="number" step="any" placeholder="Min" inputmode="decimal">
      <span class="range-sep">–</span>
      <input id="range-max" type="number" step="any" placeholder="Maks" inputmode="decimal">
      <button id="range-clear" class="range-clear" type="button" title="Bu satırı temizle" aria-label="Bu satırı temizle">×</button>
    </div>
    <button id="range-add" class="range-add" type="button" title="Bu koşulu ekle (birden fazla koşul aynı anda uygulanır)">+ Filtre ekle</button>
    <label class="switch-row">
      <span class="switch">
        <input type="checkbox" id="risk-toggle" checked>
        <span class="switch-track"><span class="switch-knob"></span></span>
      </span>
      <span class="switch-row-label">Sadece riski geçenler</span>
    </label>
    <label class="switch-row">
      <span class="switch">
        <input type="checkbox" id="veri-uyari-toggle">
        <span class="switch-track"><span class="switch-knob"></span></span>
      </span>
      <span class="switch-row-label">Sadece veri uyarısı olanlar</span>
    </label>
    <label class="switch-row" title="Favoriler bu tarayıcıda saklanır (localStorage) — başka cihaz/tarayıcıya taşınmaz.">
      <span class="switch">
        <input type="checkbox" id="favori-toggle">
        <span class="switch-track"><span class="switch-knob"></span></span>
      </span>
      <span class="switch-row-label">★ Sadece favorilerim</span>
    </label>
    <span class="result-count" id="result-count"></span>
  </section>

  <div class="filter-chips" id="filter-chips"></div>

  <div class="table-wrap">
    <table>
      <thead>
        <tr id="head-row"></tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    <div class="empty-state" id="empty-state" style="display:none;">Aramanla eşleşen fon yok.</div>
  </div>

  <footer>
    <span>Yatırım tavsiyesi değildir — sistematik bir tarama aracıdır. Karar sizindir.</span>
    <span id="fund-count-footer"></span>
  </footer>
</div>

<script>
  const DATA = __FON_DATA__;
  const RUN_DATE = "__RUN_DATE__";

  document.getElementById('run-date').textContent = RUN_DATE;
  document.getElementById('fund-count-footer').textContent = DATA.length + ' fon tarandı';

  const FAVORI_ANAHTAR = 'fonModelPortfoy_favoriler';
  let favoriler;
  try {
    favoriler = new Set(JSON.parse(localStorage.getItem(FAVORI_ANAHTAR) || '[]'));
  } catch (e) {
    favoriler = new Set();
  }
  function favorileriKaydet() {
    try {
      localStorage.setItem(FAVORI_ANAHTAR, JSON.stringify([...favoriler]));
    } catch (e) { /* ozel tarayici modu vb. - sessizce yut, favoriler bu oturumda calismaya devam eder */ }
  }

  DATA.forEach(d => {
    d.risk_passed = d.risk_status === 'GEÇTİ';
    d.veri_uyari_var = !!(d.veri_uyarilari && d.veri_uyarilari.length);
  });

  function tierOf(d) {
    if (!d.risk_passed) return 'out';
    if (d.toplam_skor == null) return 'weak';
    if (d.toplam_skor >= 90) return 'strong';
    if (d.toplam_skor >= 70) return 'watch';
    return 'weak';
  }

  function fmtPct(v, digits) {
    if (v == null) return '<span class="ret zero">—</span>';
    const cls = v > 0.05 ? 'pos' : v < -0.05 ? 'neg' : 'zero';
    const sign = v > 0 ? '+' : '';
    return `<span class="ret ${cls}">${sign}${v.toFixed(digits == null ? 1 : digits)}%</span>`;
  }
  function fmtNum(v, digits) {
    if (v == null) return '—';
    return v.toLocaleString('tr-TR', { minimumFractionDigits: digits == null ? 2 : digits, maximumFractionDigits: digits == null ? 2 : digits });
  }

  const COLUMNS = [
    { key: 'favori', label: '★', sort: (d) => favoriler.has(d.fonKodu) ? 1 : 0 },
    { key: 'fonKodu', label: 'Kod', sort: (d) => d.fonKodu },
    { key: 'fonUnvan', label: 'Fon Unvanı', sort: (d) => d.fonUnvan },
    { key: 'toplam_skor', label: 'Skor', num: true, sort: (d) => d.toplam_skor ?? -1 },
    { key: 'katman_a_getiri', label: 'Katman A (Getiri)', num: true, sort: (d) => d.katman_a_getiri ?? -1 },
    { key: 'katman_b_akis', label: 'Katman B (Akış)', num: true, sort: (d) => d.katman_b_akis ?? -1 },
    { key: 'katman_c_risk', label: 'Katman C (Risk)', num: true, sort: (d) => d.katman_c_risk ?? -1 },
    { key: 'skor_degisim_1g', label: 'Skor Değ. 1G', num: true, sort: (d) => d.skor_degisim_1g ?? -Infinity },
    { key: 'skor_degisim_1h', label: 'Skor Değ. 1H', num: true, sort: (d) => d.skor_degisim_1h ?? -Infinity },
    { key: 'getiri_1g', label: 'Günlük Getiri', num: true, sort: (d) => d.getiri_1g ?? -Infinity },
    { key: 'getiri_1h', label: 'Haftalık Getiri', num: true, sort: (d) => d.getiri_1h ?? -Infinity },
    { key: 'getiri_pct', label: '~1 Ay Getiri', num: true, sort: (d) => d.getiri_pct ?? -Infinity },
    { key: 'getiri_3a', label: '3 Ay Getiri', num: true, sort: (d) => d.getiri_3a ?? -Infinity },
    { key: 'getiri_6a', label: '6 Ay Getiri', num: true, sort: (d) => d.getiri_6a ?? -Infinity },
    { key: 'getiri_1y', label: '1 Yıl Getiri', num: true, sort: (d) => d.getiri_1y ?? -Infinity },
    { key: 'akis_oran_1g', label: 'Günlük Akış', num: true, sort: (d) => d.akis_oran_1g ?? -Infinity },
    { key: 'akis_oran_1h', label: 'Haftalık Akış', num: true, sort: (d) => d.akis_oran_1h ?? -Infinity },
    { key: 'akis_oran_pct', label: 'Akış/Büyüklük', num: true, sort: (d) => d.akis_oran_pct ?? -Infinity },
    { key: 'volatilite_1a', label: 'Volatilite (1A)', num: true, sort: (d) => d.volatilite_1a ?? -Infinity },
    { key: 'max_dusus_1a', label: 'Maks. Düşüş (1A)', num: true, sort: (d) => d.max_dusus_1a ?? -Infinity },
    { key: 'yogunlasma_bin_tl', label: 'Yoğunlaşma (Bin TL)', num: true, sort: (d) => d.yogunlasma_bin_tl ?? Infinity },
    { key: 'net_akis_mn', label: 'Net Akış (Mn TL)', num: true, sort: (d) => d.net_akis_mn ?? -Infinity },
    { key: 'portfoy_buyuklugu_mn', label: 'Büyüklük (Mn TL)', num: true, sort: (d) => d.portfoy_buyuklugu_mn ?? -Infinity },
    { key: 'kisi_sayisi', label: 'Yatırımcı', num: true, sort: (d) => d.kisi_sayisi ?? -Infinity },
    { key: 'sharpe_1a', label: 'Sharpe-Benzeri (1A)', num: true, sort: (d) => d.sharpe_1a ?? -Infinity },
    { key: 'akis_z', label: 'Akış Z-Skoru', num: true, sort: (d) => d.akis_z ?? -Infinity },
    { key: 'guncel_fiyat', label: 'Güncel Fiyat', num: true, sort: (d) => d.guncel_fiyat ?? -Infinity },
    { key: 'son_tarih', label: 'Son Veri', sort: (d) => d.son_tarih || '' },
    { key: 'veri_uyarilari', label: 'Veri Uyarısı', sort: (d) => (d.veri_uyarilari || []).length },
  ];

  // UYARI_ALAN_ESLEME: fon_model_portfoy.py'deki ayni isimli Python listesiyle
  // ELLE SENKRON tutulur - hangi uyari on-eki hangi sutun HUCRESINI vurgulayacak.
  const UYARI_ALAN_ESLEME = [
    ['Geçersiz güncel fiyat', 'guncel_fiyat'],
    ['Negatif portföy büyüklüğü', 'portfoy_buyuklugu_mn'],
    ['Negatif yatırımcı sayısı', 'kisi_sayisi'],
    ['Günlük getiri aşırı', 'getiri_1g'],
    ['~1 Ay getiri aşırı', 'getiri_pct'],
    ['Akış oranı aşırı', 'akis_oran_pct'],
    ['Negatif volatilite', 'volatilite_1a'],
    ['Veri güncel değil', 'son_tarih'],
    ['Fon kodu taramada birden fazla kez geçiyor', 'fonKodu'],
  ];

  function uyariAlanlari(d) {
    const set = new Set();
    (d.veri_uyarilari || []).forEach(u => {
      const hit = UYARI_ALAN_ESLEME.find(([onEk]) => u.startsWith(onEk));
      if (hit) set.add(hit[1]);
    });
    return set;
  }

  let sortKey = 'toplam_skor';
  let sortDir = -1;

  const headRow = document.getElementById('head-row');
  headRow.innerHTML = COLUMNS.map(c =>
    `<th data-key="${c.key}" class="${c.num ? 'num' : ''}">${c.label}<span class="arrow"></span></th>`
  ).join('');

  headRow.querySelectorAll('th[data-key]').forEach(th => {
    th.addEventListener('click', () => {
      const key = th.dataset.key;
      if (sortKey === key) { sortDir *= -1; } else { sortKey = key; sortDir = key === 'fonKodu' || key === 'fonUnvan' ? 1 : -1; }
      render();
    });
  });

  const turSel = document.getElementById('tur-filter');
  const turler = [...new Set(DATA.map(d => d.tur).filter(Boolean))].sort((a, b) => a.localeCompare(b, 'tr'));
  turSel.innerHTML = '<option value="">Tüm türler</option>' + turler.map(s => `<option value="${s}">${s}</option>`).join('');

  const rangeField = document.getElementById('range-field');
  const rangeKeySel = document.getElementById('range-key');
  const rangeMinInput = document.getElementById('range-min');
  const rangeMaxInput = document.getElementById('range-max');
  const rangeClearBtn = document.getElementById('range-clear');
  const rangeAddBtn = document.getElementById('range-add');
  const filterChipsEl = document.getElementById('filter-chips');
  const numericCols = COLUMNS.filter(c => c.num);
  rangeKeySel.innerHTML = '<option value="">Filtre tipi…</option>' +
    numericCols.map(c => `<option value="${c.key}">${c.label}</option>`).join('');

  // Birden fazla kosulu AYNI ANDA uygulayabilmek icin: "+ Filtre ekle" tiklandiginda
  // o an secili sutun/min/maks KALICI bir filtreye (aktifFiltreler) donusur ve satir
  // sifirlanir, boylece ikinci/ucuncu bir kosul daha girilebilir. Hepsi VE (AND) ile
  // birlesir. Satirdaki (henuz eklenmemis) deger de ayrica canli filtre olarak calisir.
  let aktifFiltreler = [];

  function updateRangeState() {
    const active = !!rangeKeySel.value && (rangeMinInput.value !== '' || rangeMaxInput.value !== '');
    rangeField.classList.toggle('active', active);
    rangeAddBtn.disabled = !active;
  }

  function aralikMetni(f) {
    if (f.min != null && f.max != null) return `${f.min}–${f.max}`;
    if (f.min != null) return `≥ ${f.min}`;
    return `≤ ${f.max}`;
  }

  function renderFiltreChipleri() {
    filterChipsEl.classList.toggle('visible', aktifFiltreler.length > 0);
    if (aktifFiltreler.length === 0) { filterChipsEl.innerHTML = ''; return; }
    filterChipsEl.innerHTML = aktifFiltreler.map((f, i) =>
      `<span class="filter-chip">${f.label} ${aralikMetni(f)}<button class="filter-chip-remove" data-idx="${i}" type="button" aria-label="${f.label} filtresini kaldır">×</button></span>`
    ).join('') + (aktifFiltreler.length > 1
      ? '<button class="filter-chips-clear" id="filter-chips-clear" type="button">Tümünü temizle</button>' : '');
  }

  function filtreEkle() {
    const key = rangeKeySel.value;
    const min = rangeMinInput.value === '' ? null : parseFloat(rangeMinInput.value);
    const max = rangeMaxInput.value === '' ? null : parseFloat(rangeMaxInput.value);
    if (!key || (min == null && max == null)) return;
    const col = COLUMNS.find(c => c.key === key);
    aktifFiltreler.push({ key, min, max, label: col ? col.label : key });
    rangeKeySel.value = '';
    rangeMinInput.value = '';
    rangeMaxInput.value = '';
    updateRangeState();
    renderFiltreChipleri();
    render();
  }

  [rangeKeySel, rangeMinInput, rangeMaxInput].forEach(el => {
    el.addEventListener('input', () => { updateRangeState(); render(); });
    el.addEventListener('change', () => { updateRangeState(); render(); });
  });
  rangeMinInput.addEventListener('keydown', (e) => { if (e.key === 'Enter') filtreEkle(); });
  rangeMaxInput.addEventListener('keydown', (e) => { if (e.key === 'Enter') filtreEkle(); });
  rangeAddBtn.addEventListener('click', filtreEkle);
  rangeClearBtn.addEventListener('click', () => {
    rangeKeySel.value = '';
    rangeMinInput.value = '';
    rangeMaxInput.value = '';
    updateRangeState();
    render();
  });
  filterChipsEl.addEventListener('click', (e) => {
    if (e.target.id === 'filter-chips-clear') {
      aktifFiltreler = [];
    } else {
      const btn = e.target.closest('.filter-chip-remove');
      if (!btn) return;
      aktifFiltreler.splice(parseInt(btn.dataset.idx, 10), 1);
    }
    renderFiltreChipleri();
    render();
  });
  updateRangeState();

  const searchInput = document.getElementById('search');
  const riskToggle = document.getElementById('risk-toggle');
  const veriUyariToggle = document.getElementById('veri-uyari-toggle');
  const favoriToggle = document.getElementById('favori-toggle');
  const tbody = document.getElementById('tbody');
  const emptyState = document.getElementById('empty-state');
  const resultCount = document.getElementById('result-count');

  function renderStats() {
    const passed = DATA.filter(d => d.risk_passed);
    const strong = passed.filter(d => d.toplam_skor >= 90);
    const watch = passed.filter(d => d.toplam_skor >= 70 && d.toplam_skor < 90);
    const withFlow = passed.filter(d => d.akis_oran_pct != null);
    // Aritmetik ortalama, birkac asiri uc deger (orn. pay bolunmesi kaynakli %5000+ akis)
    // tarafindan komple domine edilip anlamsizlasabiliyor - medyan uc degerlere karsi
    // dayanikli, "tipik fon" icin cok daha temsili bir ozet.
    const flowSirali = withFlow.map(d => d.akis_oran_pct).sort((a, b) => a - b);
    const n = flowSirali.length;
    const medyanFlow = n === 0 ? null : n % 2 === 1
      ? flowSirali[(n - 1) / 2]
      : (flowSirali[n / 2 - 1] + flowSirali[n / 2]) / 2;
    const top = [...passed].sort((a, b) => (b.toplam_skor ?? -1) - (a.toplam_skor ?? -1))[0];
    const uyarili = DATA.filter(d => d.veri_uyari_var);

    const tiles = [
      { label: 'Taranan Fon', value: DATA.length, sub: 'TEFAS' },
      { label: 'Riski Geçen', value: passed.length, sub: (DATA.length - passed.length) + ' elendi' },
      { label: 'Öne Çıkan', value: strong.length, sub: 'skor ≥ 90' },
      { label: 'İzlemede', value: watch.length, sub: 'skor 70–89' },
      { label: 'Medyan Akış/Büyüklük', value: medyanFlow == null ? '—' : (medyanFlow > 0 ? '+' : '') + medyanFlow.toFixed(1) + '%', sub: '~30 günlük, tipik fon', cls: medyanFlow > 0 ? 'pos' : medyanFlow < 0 ? 'neg' : '' },
      { label: 'En Yüksek Skor', value: top ? top.fonKodu : '—', sub: top ? top.toplam_skor.toFixed(1) + ' puan' : '' },
      { label: 'Veri Uyarısı', value: uyarili.length, sub: uyarili.length ? 'kontrol et' : 'temiz', cls: uyarili.length ? 'neg' : 'pos' },
      { label: 'Favorilerim', value: favoriler.size, sub: favoriler.size ? '★ ile gör' : 'yıldızla ekle' },
    ];
    document.getElementById('stats').innerHTML = tiles.map(t =>
      `<div class="stat"><span class="stat-label">${t.label}</span><span class="stat-value ${t.cls || ''}">${t.value}</span><span class="stat-sub">${t.sub}</span></div>`
    ).join('');
  }

  function currentFilter() {
    const q = searchInput.value.trim().toLocaleLowerCase('tr');
    const tur = turSel.value;
    const onlyPassed = riskToggle.checked;
    const onlyFlagged = veriUyariToggle.checked;
    const onlyFavori = favoriToggle.checked;
    const rangeKey = rangeKeySel.value;
    const rangeMin = rangeMinInput.value === '' ? null : parseFloat(rangeMinInput.value);
    const rangeMax = rangeMaxInput.value === '' ? null : parseFloat(rangeMaxInput.value);
    return DATA.filter(d => {
      if (onlyPassed && !d.risk_passed) return false;
      if (onlyFlagged && !d.veri_uyari_var) return false;
      if (onlyFavori && !favoriler.has(d.fonKodu)) return false;
      if (tur && d.tur !== tur) return false;
      if (q && !(d.fonKodu.toLocaleLowerCase('tr').includes(q) || (d.fonUnvan || '').toLocaleLowerCase('tr').includes(q))) return false;
      if (rangeKey && (rangeMin != null || rangeMax != null)) {
        const val = d[rangeKey];
        if (val == null) return false;
        if (rangeMin != null && val < rangeMin) return false;
        if (rangeMax != null && val > rangeMax) return false;
      }
      for (const f of aktifFiltreler) {
        const val = d[f.key];
        if (val == null) return false;
        if (f.min != null && val < f.min) return false;
        if (f.max != null && val > f.max) return false;
      }
      return true;
    });
  }

  function render() {
    headRow.querySelectorAll('th[data-key]').forEach(th => {
      th.classList.toggle('active', th.dataset.key === sortKey);
      th.querySelector('.arrow').textContent = th.dataset.key === sortKey ? (sortDir === 1 ? '▲' : '▼') : '';
    });

    const col = COLUMNS.find(c => c.key === sortKey);
    const rows = currentFilter().sort((a, b) => {
      const av = col.sort(a), bv = col.sort(b);
      if (typeof av === 'string') return av.localeCompare(bv, 'tr') * sortDir;
      return (av - bv) * sortDir;
    });

    resultCount.textContent = rows.length + ' / ' + DATA.length + ' fon';
    emptyState.style.display = rows.length ? 'none' : 'block';

    tbody.innerHTML = rows.map(d => {
      const tier = tierOf(d);
      const uyariAlan = uyariAlanlari(d);
      const wc = (key) => uyariAlan.has(key) ? ' cell-warn' : '';
      const favoriDurum = favoriler.has(d.fonKodu);
      return `
      <tr class="row ${tier}">
        <td class="fav-cell"><button class="fav-star ${favoriDurum ? 'active' : ''}" data-kod="${d.fonKodu}" title="${favoriDurum ? 'Favorilerden çıkar' : 'Favorilere ekle'}" aria-label="${favoriDurum ? 'Favorilerden çıkar' : 'Favorilere ekle'}">${favoriDurum ? '★' : '☆'}</button></td>
        <td class="ticker${wc('fonKodu')}">${d.fonKodu}</td>
        <td class="name">${d.fonUnvan || ''}<span class="sector">${d.tur || ''}</span></td>
        <td class="num"><div class="score-cell"><span class="score-num">${d.toplam_skor != null ? d.toplam_skor.toFixed(1) : '—'}</span><span class="score-track"><span class="score-fill" style="width:${Math.max(0, Math.min(100, d.toplam_skor ?? 0))}%"></span></span></div></td>
        <td class="num">${d.katman_a_getiri != null ? d.katman_a_getiri.toFixed(1) : '—'}</td>
        <td class="num">${d.katman_b_akis != null ? d.katman_b_akis.toFixed(1) : '—'}</td>
        <td class="num">${d.katman_c_risk != null ? d.katman_c_risk.toFixed(1) : '—'}</td>
        <td class="num">${fmtPct(d.skor_degisim_1g)}</td>
        <td class="num">${fmtPct(d.skor_degisim_1h)}</td>
        <td class="num${wc('getiri_1g')}">${fmtPct(d.getiri_1g)}</td>
        <td class="num">${fmtPct(d.getiri_1h)}</td>
        <td class="num${wc('getiri_pct')}">${fmtPct(d.getiri_pct)}</td>
        <td class="num">${fmtPct(d.getiri_3a)}</td>
        <td class="num">${fmtPct(d.getiri_6a)}</td>
        <td class="num">${fmtPct(d.getiri_1y)}</td>
        <td class="num">${fmtPct(d.akis_oran_1g)}</td>
        <td class="num">${fmtPct(d.akis_oran_1h)}</td>
        <td class="num${wc('akis_oran_pct')}">${fmtPct(d.akis_oran_pct)}</td>
        <td class="num${wc('volatilite_1a')}">${fmtNum(d.volatilite_1a, 2)}</td>
        <td class="num">${fmtPct(d.max_dusus_1a)}</td>
        <td class="num">${fmtNum(d.yogunlasma_bin_tl, 1)}</td>
        <td class="num">${fmtNum(d.net_akis_mn, 1)}</td>
        <td class="num${wc('portfoy_buyuklugu_mn')}">${fmtNum(d.portfoy_buyuklugu_mn, 0)}</td>
        <td class="num${wc('kisi_sayisi')}">${d.kisi_sayisi != null ? d.kisi_sayisi.toLocaleString('tr-TR') : '—'}</td>
        <td class="num">${fmtNum(d.sharpe_1a, 2)}</td>
        <td class="num">${fmtNum(d.akis_z, 2)}</td>
        <td class="num${wc('guncel_fiyat')}">${fmtNum(d.guncel_fiyat, 4)}</td>
        <td class="num${wc('son_tarih')}">${d.son_tarih || '—'}</td>
        <td class="warn-cell">${d.veri_uyari_var ? `<span class="chip out" title="${(d.veri_uyarilari || []).join('; ').replace(/"/g, '&quot;')}">⚠ ${d.veri_uyarilari.length}</span>` : '—'}</td>
      </tr>`;
    }).join('');
  }

  searchInput.addEventListener('input', render);
  turSel.addEventListener('change', render);
  riskToggle.addEventListener('change', render);
  veriUyariToggle.addEventListener('change', render);
  favoriToggle.addEventListener('change', render);

  tbody.addEventListener('click', (e) => {
    const btn = e.target.closest('.fav-star');
    if (!btn) return;
    const kod = btn.dataset.kod;
    if (favoriler.has(kod)) favoriler.delete(kod); else favoriler.add(kod);
    favorileriKaydet();
    render();
    renderStats();
  });

  renderStats();
  render();
</script>
</body>
</html>
"""


def write_html_dashboard(rows, run_date_str, output_path: Path):
    """Skor + akis verisini arama/filtre/siralama ozellikli, tek dosyalik bagimsiz
    bir HTML sayfasina yazar."""
    dashboard_rows = []
    for r in rows:
        dashboard_rows.append({
            "fonKodu": r["fonKodu"], "fonUnvan": r["fonUnvan"], "tur": r["tur_aciklama"],
            "toplam_skor": r["toplam_skor"],
            "katman_a_getiri": r.get("katman_a_getiri"), "katman_b_akis": r.get("katman_b_akis"),
            "katman_c_risk": r.get("katman_c_risk"),
            "skor_degisim_1g": r.get("skor_degisim_1g"), "skor_degisim_1h": r.get("skor_degisim_1h"),
            "getiri_1g": r.get("getiri_1g"), "getiri_1h": r.get("getiri_1h"), "getiri_pct": r["getiri_pct"],
            "getiri_3a": r.get("getiri_3a"), "getiri_6a": r.get("getiri_6a"), "getiri_1y": r.get("getiri_1y"),
            "akis_oran_1g": r.get("akis_oran_1g"), "akis_oran_1h": r.get("akis_oran_1h"),
            "akis_oran_pct": r["akis_oran_pct"],
            "volatilite_1a": r.get("volatilite_1a"), "max_dusus_1a": r.get("max_dusus_1a"),
            "sharpe_1a": r.get("sharpe_1a"), "akis_z": r.get("akis_z"),
            "yogunlasma_bin_tl": (r["yogunlasma_tl"] / 1_000) if r.get("yogunlasma_tl") is not None else None,
            "net_akis_mn": (r["net_akis_tl"] / 1_000_000) if r["net_akis_tl"] is not None else None,
            "portfoy_buyuklugu_mn": (r["guncel_portfoy_buyuklugu"] / 1_000_000) if r["guncel_portfoy_buyuklugu"] else None,
            "kisi_sayisi": r["guncel_kisi_sayisi"],
            "risk_status": r["risk_status"],
            "guncel_fiyat": r.get("guncel_fiyat"),
            "son_tarih": r.get("son_tarih"),
            "veri_uyarilari": r.get("veri_uyarilari") or [],
        })
    data_json = json.dumps(dashboard_rows, ensure_ascii=False)
    html = DASHBOARD_TEMPLATE.replace("__FON_DATA__", data_json).replace("__RUN_DATE__", run_date_str)
    output_path.write_text(html, encoding="utf-8")
    print(f"HTML panosu yazildi: {output_path}")


def git_publish(run_date_str):
    """docs/fon.html'i GitHub'a commit+push eder, boylece GitHub Pages uzerindeki
    pano otomatik guncellenir. Repo/remote/kimlik bilgisi eksikse veya push
    basarisiz olursa scripti durdurmadan uyari basar."""
    import subprocess
    repo_dir = Path(__file__).resolve().parent
    candidate_files = ["docs/fon.html", "fon_skor_gecmis.csv", "fon_akis_gecmis.csv"]
    files_to_add = [f for f in candidate_files if (repo_dir / f).exists()]
    try:
        subprocess.run(
            ["git", "add"] + files_to_add,
            cwd=repo_dir, check=True, capture_output=True, text=True,
        )
        commit = subprocess.run(
            ["git", "commit", "-m", f"Fon taramasi guncellemesi - {run_date_str}"],
            cwd=repo_dir, capture_output=True, text=True,
        )
        if commit.returncode != 0:
            if "nothing to commit" in commit.stdout:
                print("  Git: degisiklik yok, commit atlandi.")
            else:
                print(f"  [uyari] git commit basarisiz: {commit.stdout}{commit.stderr}")
            return
        push = subprocess.run(
            ["git", "push", "origin", "main"],
            cwd=repo_dir, capture_output=True, text=True,
        )
        if push.returncode != 0:
            print(f"  [uyari] git push basarisiz (ilk push'u elle yapman gerekebilir): {push.stderr.strip()}")
        else:
            print("  GitHub'a push edildi - Pages birkac dakika icinde guncellenir.")
    except Exception as e:
        print(f"  [uyari] Otomatik git publish basarisiz: {e}")


# =============================================================================
# EXCEL RAPORU
# =============================================================================

def write_excel_report(rows, run_date_str, output_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    wb = Workbook()
    ws = wb.active
    ws.title = "Fon Model Portföy"
    ws.sheet_view.showGridLines = False

    title_font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    header_font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    normal_font = Font(name=FONT, size=10)
    warn_font = Font(name=FONT, size=9, bold=True, color="FFFFFF")

    fill_navy = PatternFill("solid", fgColor="1F3864")
    fill_red = PatternFill("solid", fgColor="C00000")
    fill_header = PatternFill("solid", fgColor="2E75B6")

    thin = Side(style="thin", color="BFBFBF")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["Sıra", "Kod", "Fon Unvanı", "Tür", "Toplam Skor",
               "Katman A (Getiri)", "Katman B (Akış)", "Katman C (Risk)",
               "Skor Değ. (1G)", "Skor Değ. (1H)",
               "Günlük Getiri %", "Haftalık Getiri %", "~1 Ay Getiri %",
               "3 Ay Getiri %", "6 Ay Getiri %", "1 Yıl Getiri %",
               "Günlük Akış %", "Haftalık Akış %", "Akış/Büyüklük % (1A)",
               "Volatilite % (1A)", "Maks. Düşüş % (1A)", "Yoğunlaşma (Bin TL/Kişi)",
               "Net Akış (Mn TL)", "Portföy Büyüklüğü (Mn TL)", "Yatırımcı Sayısı", "Risk Filtresi",
               "Sharpe-Benzeri (1A)", "Akış Z-Skoru (Kendi Geçmişi)", "Güncel Fiyat", "Son Veri Tarihi",
               "Veri Uyarısı"]
    widths = [6, 8, 36, 17, 11, 13, 12, 12, 11, 11, 10, 11, 11, 10, 10, 10, 11, 11, 13, 12, 13, 15, 13, 16, 12, 20,
              14, 16, 12, 13, 40]
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}
    fill_uyari_hucre = PatternFill("solid", fgColor="FFE699")
    last_col = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"TEFAS FON MODEL PORTFÖY — Tarama ({run_date_str})"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = ("UYARI: Yatırım tavsiyesi değildir, sistematik bir tarama aracıdır. "
                "Veri kaynağı: TEFAS (tefas.gov.tr). Karar sizindir.")
    ws["A2"].font = warn_font
    ws["A2"].fill = fill_red
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 20

    for i, (h, w) in enumerate(zip(headers, widths)):
        col = get_column_letter(i + 1)
        ws.column_dimensions[col].width = w
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = header_font
        c.fill = fill_header
        c.alignment = center
        c.border = border_all
    ws.row_dimensions[4].height = 32

    def sort_key(r):
        if not r["risk_passed"]:
            return (2, 0)
        if r["toplam_skor"] is None:
            return (1, 0)
        return (0, -r["toplam_skor"])

    rows_sorted = sorted(rows, key=sort_key)

    r_idx = 5
    for i, r in enumerate(rows_sorted):
        vals = [
            i + 1, r["fonKodu"], r["fonUnvan"], r["tur_aciklama"],
            round(r["toplam_skor"], 1) if r["toplam_skor"] is not None else None,
            round(r["katman_a_getiri"], 1) if r.get("katman_a_getiri") is not None else None,
            round(r["katman_b_akis"], 1) if r.get("katman_b_akis") is not None else None,
            round(r["katman_c_risk"], 1) if r.get("katman_c_risk") is not None else None,
            round(r["skor_degisim_1g"], 1) if r.get("skor_degisim_1g") is not None else None,
            round(r["skor_degisim_1h"], 1) if r.get("skor_degisim_1h") is not None else None,
            round(r["getiri_1g"], 2) if r.get("getiri_1g") is not None else None,
            round(r["getiri_1h"], 2) if r.get("getiri_1h") is not None else None,
            round(r["getiri_pct"], 2) if r["getiri_pct"] is not None else None,
            round(r["getiri_3a"], 2) if r.get("getiri_3a") is not None else None,
            round(r["getiri_6a"], 2) if r.get("getiri_6a") is not None else None,
            round(r["getiri_1y"], 2) if r.get("getiri_1y") is not None else None,
            round(r["akis_oran_1g"], 2) if r.get("akis_oran_1g") is not None else None,
            round(r["akis_oran_1h"], 2) if r.get("akis_oran_1h") is not None else None,
            round(r["akis_oran_pct"], 2) if r["akis_oran_pct"] is not None else None,
            round(r["volatilite_1a"], 2) if r.get("volatilite_1a") is not None else None,
            round(r["max_dusus_1a"], 2) if r.get("max_dusus_1a") is not None else None,
            round(r["yogunlasma_tl"] / 1_000, 1) if r.get("yogunlasma_tl") is not None else None,
            round(r["net_akis_tl"] / 1_000_000, 2) if r["net_akis_tl"] is not None else None,
            round(r["guncel_portfoy_buyuklugu"] / 1_000_000, 1) if r["guncel_portfoy_buyuklugu"] else None,
            r["guncel_kisi_sayisi"],
            r["risk_status"],
            round(r["sharpe_1a"], 2) if r.get("sharpe_1a") is not None else None,
            round(r["akis_z"], 2) if r.get("akis_z") is not None else None,
            round(r["guncel_fiyat"], 6) if r.get("guncel_fiyat") is not None else None,
            r.get("son_tarih"),
            "; ".join(r.get("veri_uyarilari") or []) or "—",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=col, value=v)
            cell.font = normal_font
            cell.border = border_all
            cell.alignment = left if col in (3, 26, len(headers)) else center

        # Veri Uyarisi metnini okumaya gerek kalmadan, SORUNUN OLDUGU HUCRENIN
        # kendisini vurgula (bkz. UYARI_ALAN_ESLEME) - "nerede" sorusuna Excel
        # icinde dogrudan cevap verir.
        for uyari in (r.get("veri_uyarilari") or []):
            for on_ek, sutun_adi, _ in UYARI_ALAN_ESLEME:
                if uyari.startswith(on_ek):
                    hedef_col = header_to_col.get(sutun_adi)
                    if hedef_col:
                        ws.cell(row=r_idx, column=hedef_col).fill = fill_uyari_hucre
                    break
        r_idx += 1

    last_row = r_idx - 1
    if last_row >= 5:
        ws.conditional_formatting.add(
            f"E5:E{last_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["90"],
                       fill=PatternFill("solid", fgColor="8EA9DB")))
        ws.conditional_formatting.add(
            f"E5:E{last_row}",
            CellIsRule(operator="between", formula=["70", "89.9"],
                       fill=PatternFill("solid", fgColor="A9D18E")))
        for col_letter in ("I", "J"):
            ws.conditional_formatting.add(
                f"{col_letter}5:{col_letter}{last_row}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="A9D18E")))
            ws.conditional_formatting.add(
                f"{col_letter}5:{col_letter}{last_row}",
                CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FF7C80")))
        ws.conditional_formatting.add(
            f"Z5:Z{last_row}",
            CellIsRule(operator="notEqual", formula=['"GEÇTİ"'],
                       fill=PatternFill("solid", fgColor="FF7C80")))
        veri_uyari_col = get_column_letter(len(headers))
        ws.conditional_formatting.add(
            f"{veri_uyari_col}5:{veri_uyari_col}{last_row}",
            CellIsRule(operator="notEqual", formula=['"—"'],
                       fill=PatternFill("solid", fgColor="FFE699")))

    ws.freeze_panes = "A5"

    note_row = last_row + 3
    ws.merge_cells(f"A{note_row}:{get_column_letter(len(headers))}{note_row}")
    ws[f"A{note_row}"] = (
        "Not: Toplam Skor = ağırlıklı ortalama(Katman A %45, Katman B %20, Katman C %35), sonra kendi "
        "türü içinde tekrar yüzdelik dilime çevrilir (dağılımı düz tutmak için). Ağırlıklar, risk-ayarlı "
        "getirinin akış sinyalinden daha kalıcı olduğu bulgusuna dayanır (Morningstar/Lipper metodolojileri "
        "ve akademik literatür). Katman A = 1 Hafta/1 Ay/3 Ay/6 Ay/1 Yıl getirileri + Sharpe-benzeri oranın "
        "(1 Ay getiri / 1 Ay volatilite) ortalama percentile'ı. Katman B = 1 Hafta/1 Ay akış oranı + Akış "
        "Z-Skoru'nun ortalama percentile'ı. Katman C = düşük volatilite + küçük maks. düşüş + düşük "
        "yatırımcı yoğunlaşmasının ortalama percentile'ı (üçü de \"düşük risk\" yönünde normalize "
        "edilmiştir). Sharpe-Benzeri (1A), risksiz faiz oranı düşülmemiş/yıllandırılmamış basit bir "
        "getiri/risk oranıdır — sadece AYNI TÜR içindeki fonları sıralamak için kullanılır, mutlak değeri "
        "yorumlanmamalıdır. Akış Z-Skoru, fonun bugünkü akış/büyüklük oranının KENDİ geçmiş ortalamasına "
        "göre kaç standart sapma uzakta olduğunu gösterir (pozitif = normalden fazla para giriyor); en az "
        f"{AKIS_Z_MIN_GOZLEM} günlük geçmiş birikene kadar boş gelir. Skor Değ. (1G/1H), Toplam Skor'un "
        "bir/yedi gün önceki değerine göre değişimidir — ilk birkaç çalıştırmada boş gelir. Net Akış, "
        "tedavüldeki pay sayısı değişiminden tahmin edilir. ~1 Ay/3 Ay/6 Ay/1 Yıl getirileri, TEFAS'ın "
        "kendi \"Getiri Bilgisi\" hesabıyla BİREBİR eşleşecek şekilde hesaplanır: bugünden tam 1/3/6/12 "
        "takvim ayı önceki AYNI GÜN referans alınır (kısa aylarda ayın son gününe sabitlenir), o gün işlem "
        "günü değilse EN YAKIN SONRAKI işlem gününün fiyatı kullanılır (TEFAS'ın kendi kuralı — geriye "
        "değil ileriye yuvarlanır). Veri Uyarısı, her çalıştırmada otomatik "
        "çalışan bir mantık kontrolüdür (negatif/imkânsız değerler, aşırı büyük hareketler, bayat veri, "
        "tekrar eden fon kodu) — skorlamayı etkilemez, sadece altındaki HAM verinin makul olup olmadığını "
        "işaretler; boş (—) olması verinin kesin doğru olduğunu garanti etmez, sadece bariz hataları eler. "
        "Bir uyarı tetiklendiğinde, o uyarının kaynağı olan HÜCRE (örn. aşırı akış → Akış/Büyüklük % "
        "hücresi, bayat veri → Son Veri Tarihi hücresi) sarı renkle ayrıca vurgulanır — metni okumadan "
        "\"nerede\" sorusuna doğrudan cevap verir. Güncel Fiyat ve Son Veri Tarihi sütunları sadece bu "
        "kontrol için eklenmiştir, başka bir hesaplamada kullanılmaz."
    )
    ws[f"A{note_row}"].font = Font(name=FONT, size=9, italic=True, color="7F7F7F")
    ws[f"A{note_row}"].alignment = left

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel raporu yazildi: {output_path}")


# =============================================================================
# ANA AKIŞ
# =============================================================================

def main():
    run_date = date.today()
    run_date_str = run_date.isoformat()
    print(f"=== TEFAS Fon Model Portföy Taraması — {run_date_str} ===\n")

    print("Şemsiye fon türleri çekiliyor...")
    fon_turleri = fetch_fon_turleri()
    print(f"  -> {len(fon_turleri)} tür bulundu.\n")

    # LOOKBACK_DAYS tam 30 gun geriye gider, ama "30 gun once" bir haftasonuna/tatile denk
    # gelirse (o gun islem yoksa) ~1 Ay Getiri icin referans noktasi bulunamaz ve kod sessizce
    # serideki EN ESKI noktaya (birkac gun daha kisa bir pencereye) duser. TEFAS API'nin izin
    # verdigi ust sinir (31 gun) kadar +1 gun ekstra tampon cekerek bu durumun COGUNU (tek
    # gunluk haftasonu/tatil komsulugu) onceden onluyoruz - riski/volatiliteyi etkilemez,
    # cunku _pencere_risk_metrikleri zaten sadece hedef tarihten SONRAKI gunleri kullanir.
    FETCH_TAMPON_GUN = 1
    bas_tarih = run_date - timedelta(days=LOOKBACK_DAYS + FETCH_TAMPON_GUN)
    print(f"Tüm fonların {bas_tarih} - {run_date} arası zaman serisi çekiliyor (tek istek, biraz sürebilir)...")
    zaman_serisi, toplam_satir = fetch_tum_fonlar_zaman_serisi(bas_tarih, run_date)
    print(f"  -> {len(zaman_serisi)} fon, {toplam_satir} satır çekildi.\n")

    print("Metrikler hesaplanıyor...")
    ham_sonuclar = []
    for fon_kodu, series in zaman_serisi.items():
        metrics = compute_fund_metrics(series)
        if metrics is None:
            continue
        sfon_tur = guess_fon_turu(metrics.get("fonUnvan"))
        risk_passed, risk_status = apply_risk_filter(metrics)
        ham_sonuclar.append({
            "fonKodu": fon_kodu,
            "sfon_tur": sfon_tur,
            "tur_aciklama": fon_turleri.get(sfon_tur, "Bilinmiyor"),
            "risk_passed": risk_passed,
            "risk_status": risk_status,
            **metrics,
        })

    print(f"  -> {len(ham_sonuclar)} fon için metrik hesaplandı.\n")

    print("Veri kalitesi kontrol ediliyor...")
    veri_uyarilari = validate_data_quality(ham_sonuclar, run_date)
    if veri_uyarilari:
        print(f"  -> {len(veri_uyarilari)} fonda veri uyarısı bulundu (detaylar Excel/panonun son sütununda):")
        for u in veri_uyarilari[:5]:
            print(f"     {u['fonKodu']}: {'; '.join(u['uyarilar'])}")
        if len(veri_uyarilari) > 5:
            print(f"     ... ve {len(veri_uyarilari) - 5} fon daha.")
    else:
        print("  -> Uyarı yok, tüm veriler makul aralıkta.")
    print()

    print("Referans fiyatlar çekiliyor (1/3/6/12 ay — TEFAS'ın kendi 'Getiri Bilgisi' hesabıyla "
          "BİREBİR eşleşecek şekilde takvim ayı + ileri yuvarlama)...")
    referans_fiyatlar = {}
    for etiket, ay_sayisi in UZUN_VADE_DONEMLERI:
        hedef = n_ay_once(run_date, ay_sayisi)
        referans_fiyatlar[etiket] = fetch_referans_fiyatlar(hedef)
        print(f"  {etiket} ({hedef}): {len(referans_fiyatlar[etiket])} fon için referans fiyat bulundu.")
        time.sleep(1.0)

    for r in ham_sonuclar:
        for etiket, _ in UZUN_VADE_DONEMLERI:
            ref_fiyat = referans_fiyatlar[etiket].get(r["fonKodu"])
            getiri = (
                (r["guncel_fiyat"] - ref_fiyat) / ref_fiyat * 100.0
                if ref_fiyat else None
            )
            # "1a" -> getiri_pct alanini TEFAS-birebir degerle EZER (compute_fund_metrics'teki
            # LOOKBACK_DAYS-bazli tahmin sadece Akis/Buyukluk hesabinda kullanilmaya devam eder).
            alan_adi = "getiri_pct" if etiket == "1a" else f"getiri_{etiket}"
            r[alan_adi] = getiri
    print()

    print("Akış z-skoru (fonun kendi geçmişine göre) hesaplanıyor...")
    akis_gecmisi = load_fon_akis_gecmis()
    for r in ham_sonuclar:
        gecmis_for_fon = akis_gecmisi.get(r["fonKodu"], [])
        r["akis_z"] = compute_akis_z(gecmis_for_fon, r.get("akis_oran_pct"))
    append_fon_akis_gecmis(ham_sonuclar, run_date)
    print()

    print("Tür-bazlı yüzdelik dilim (percentile) skorları hesaplanıyor...")

    # Katman A (Getiri, coklu donem + sharpe-benzeri risk-ayarli oran), Katman B
    # (Akis, coklu donem + kendi gecmisine gore z-skor) ve Katman C (Risk: dusuk
    # volatilite + kucuk maks. dusus + dusuk yogunlasma = iyi) icin her metrigin
    # kendi turu icindeki dagilimini kuruyoruz.
    GETIRI_ALANLARI = ["getiri_1h", "getiri_pct", "getiri_3a", "getiri_6a", "getiri_1y", "sharpe_1a"]  # getiri_pct = ~1a
    AKIS_ALANLARI = ["akis_oran_1h", "akis_oran_pct", "akis_z"]  # akis_oran_pct = ~1a
    # Risk alanlarinda "iyi" yon farkli: volatilite/yogunlasma dusuk=iyi (ters cevrilecek),
    # max_dusus zaten buyuk (0'a yakin, daha az negatif) = iyi (dogrudan kullanilabilir).
    RISK_TERS_ALANLARI = ["volatilite_1a", "yogunlasma_tl"]      # dusuk deger iyi -> percentile ters cevrilir
    RISK_DUZ_ALANLARI = ["max_dusus_1a"]                          # buyuk (az negatif) deger iyi -> dogrudan

    dagilim_by_tur = defaultdict(lambda: defaultdict(list))
    for r in ham_sonuclar:
        if not r["risk_passed"]:
            continue
        for alan in GETIRI_ALANLARI + AKIS_ALANLARI + RISK_TERS_ALANLARI + RISK_DUZ_ALANLARI:
            if r.get(alan) is not None:
                dagilim_by_tur[r["sfon_tur"]][alan].append(r[alan])
    for tur in dagilim_by_tur:
        for alan in dagilim_by_tur[tur]:
            dagilim_by_tur[tur][alan].sort()

    def perc(r, alan):
        return percentile_rank(r.get(alan), dagilim_by_tur.get(r["sfon_tur"], {}).get(alan, []))

    for r in ham_sonuclar:
        if not r["risk_passed"]:
            r["katman_a_getiri"] = None
            r["katman_b_akis"] = None
            r["katman_c_risk"] = None
            r["toplam_skor_ham"] = None
            continue

        getiri_pcts = [p for p in (perc(r, a) for a in GETIRI_ALANLARI) if p is not None]
        r["katman_a_getiri"] = sum(getiri_pcts) / len(getiri_pcts) if getiri_pcts else None

        akis_pcts = [p for p in (perc(r, a) for a in AKIS_ALANLARI) if p is not None]
        r["katman_b_akis"] = sum(akis_pcts) / len(akis_pcts) if akis_pcts else None

        risk_pcts = []
        for a in RISK_TERS_ALANLARI:
            p = perc(r, a)
            if p is not None:
                risk_pcts.append(100.0 - p)   # dusuk deger iyiydi -> ters cevir
        for a in RISK_DUZ_ALANLARI:
            p = perc(r, a)
            if p is not None:
                risk_pcts.append(p)
        r["katman_c_risk"] = sum(risk_pcts) / len(risk_pcts) if risk_pcts else None

        # Esit agirlik yerine KATMAN_AGIRLIKLARI kullanilir (bkz. ayarlar bolumu).
        # Eksik katman varsa (orn. yeni fonda akis gecmisi yok), kalan katmanlarin
        # agirliklari kendi aralarinda yeniden normalize edilir.
        katman_degerleri = {"katman_a_getiri": r["katman_a_getiri"],
                             "katman_b_akis": r["katman_b_akis"],
                             "katman_c_risk": r["katman_c_risk"]}
        toplam_agirlik = sum(KATMAN_AGIRLIKLARI[k] for k, v in katman_degerleri.items() if v is not None)
        if toplam_agirlik > 0:
            r["toplam_skor_ham"] = sum(KATMAN_AGIRLIKLARI[k] * v for k, v in katman_degerleri.items()
                                        if v is not None) / toplam_agirlik
        else:
            r["toplam_skor_ham"] = None

    # Iki bagimsiz yuzdelik dilimin ortalamasi istatistiksel olarak ortaya yigilir
    # (iki zar atip toplaminin 7'ye yigilmasi gibi) - fonlarin cogu 40-60 bandinda
    # kumelenir, secim zorlasir. Bunu duzeltmek icin ham toplam skoru KENDI TURU
    # icinde TEKRAR yuzdelik dilime ceviriyoruz ("percentile of percentile") -
    # goreceli siralama aynen korunur ama dagilim zorla duzgun 0-100'e yayilir.
    ham_skor_by_tur = defaultdict(list)
    for r in ham_sonuclar:
        if r["risk_passed"] and r["toplam_skor_ham"] is not None:
            ham_skor_by_tur[r["sfon_tur"]].append(r["toplam_skor_ham"])
    for tur in ham_skor_by_tur:
        ham_skor_by_tur[tur].sort()

    for r in ham_sonuclar:
        if not r["risk_passed"] or r["toplam_skor_ham"] is None:
            r["toplam_skor"] = None
            continue
        r["toplam_skor"] = percentile_rank(r["toplam_skor_ham"], ham_skor_by_tur.get(r["sfon_tur"], []))

    print("Skor değişimi (günlük/haftalık) hesaplanıyor...")
    skor_gecmisi = load_fon_skor_gecmis()
    for r in ham_sonuclar:
        gecmis_for_fon = skor_gecmisi.get(r["fonKodu"], [])
        r["skor_degisim_1g"] = compute_skor_degisim(gecmis_for_fon, r["toplam_skor"], run_date, 1)
        r["skor_degisim_1h"] = compute_skor_degisim(gecmis_for_fon, r["toplam_skor"], run_date, 7)
    append_fon_skor_gecmis(ham_sonuclar, run_date)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"TEFAS_Fon_Model_Portfoy_{run_date_str.replace('-', '')}.xlsx"
    write_excel_report(ham_sonuclar, run_date_str, out_path)

    html_path = OUTPUT_DIR / f"TEFAS_Fon_Model_Portfoy_{run_date_str.replace('-', '')}.html"
    write_html_dashboard(ham_sonuclar, run_date_str, html_path)

    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    docs_fon = DOCS_DIR / "fon.html"
    write_html_dashboard(ham_sonuclar, run_date_str, docs_fon)

    print("\n=== Tamamlandı ===")
    print(f"Excel raporu: {out_path}")
    print(f"HTML panosu: {html_path}")

    if GIT_AUTO_PUBLISH:
        print("\nGitHub'a yayinlaniyor...")
        git_publish(run_date_str)


if __name__ == "__main__":
    main()
